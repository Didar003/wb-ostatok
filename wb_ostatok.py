import streamlit as st
import streamlit.components.v1 as components
import requests
import pandas as pd
from datetime import datetime, timedelta, date
import io
import time
import json
import os
import base64
import zipfile
import hashlib
from PIL import Image

st.set_page_config(page_title="Wildberries Отчёт", page_icon="📦", layout="wide")
st.markdown("<style>.block-container{padding-top:1.5rem;}</style>", unsafe_allow_html=True)

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "wb_data")
os.makedirs(DATA_DIR, exist_ok=True)

FBO_FILE      = os.path.join(DATA_DIR, "fbo_data.json")
SEBEST_FILE   = os.path.join(DATA_DIR, "sebest_data.json")
LOGISTIC_FILE = os.path.join(DATA_DIR, "logistic_data.json")

PERSIST_FILES = ("sebest_data.json", "fbo_data.json", "logistic_data.json")

MARKETPLACE_BASE = "https://marketplace-api.wildberries.ru"

def _tmp(name):
    return os.path.join(DATA_DIR, name)

def _gh_headers():
    token = st.secrets.get("GITHUB_TOKEN", "")
    if not token:
        return None
    return {"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json"}

def _gh_repo():
    return st.secrets.get("GITHUB_REPO", "")

def _gh_branch():
    return st.secrets.get("GITHUB_BRANCH", "main")

def github_load(filename):
    headers = _gh_headers()
    repo    = _gh_repo()
    if not headers or not repo:
        return None
    try:
        url = f"https://api.github.com/repos/{repo}/contents/wb_data/{filename}"
        r = requests.get(url, headers=headers, params={"ref": _gh_branch()}, timeout=10)
        if r.status_code == 200:
            content = base64.b64decode(r.json()["content"]).decode("utf-8")
            return json.loads(content)
    except:
        pass
    return None

def github_save(filename, data):
    headers = _gh_headers()
    repo    = _gh_repo()
    if not headers or not repo:
        return False
    try:
        url = f"https://api.github.com/repos/{repo}/contents/wb_data/{filename}"
        sha = None
        r = requests.get(url, headers=headers, params={"ref": _gh_branch()}, timeout=10)
        if r.status_code == 200:
            sha = r.json().get("sha")
        content_b64 = base64.b64encode(
            json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
        ).decode("utf-8")
        payload = {
            "message": f"auto: update {filename}",
            "content": content_b64,
            "branch": _gh_branch(),
        }
        if sha:
            payload["sha"] = sha
        r2 = requests.put(url, headers=headers, json=payload, timeout=15)
        return r2.status_code in (200, 201)
    except:
        return False

def load_json(path):
    filename = os.path.basename(path)
    cache_key = f"_json_cache_{filename}"
    if cache_key in st.session_state:
        return st.session_state[cache_key]
    data = None
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
    except:
        pass
    if data is None and filename in PERSIST_FILES:
        gh_data = github_load(filename)
        if gh_data is not None:
            data = gh_data
            try:
                os.makedirs(os.path.dirname(path), exist_ok=True)
                with open(path, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
            except:
                pass
    if data is None:
        data = {}
    st.session_state[cache_key] = data
    return data

def save_json(path, data):
    filename = os.path.basename(path)
    cache_key = f"_json_cache_{filename}"
    st.session_state[cache_key] = data
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        st.warning(f"Не сохранено локально: {e}")
    if filename in PERSIST_FILES:
        github_save(filename, data)

def _auth_token(secret_value, kind):
    """Парольден қайтарылмайтын қысқа токен жасайды (URL-де сақтау үшін)."""
    return hashlib.sha256(f"{kind}:{secret_value}".encode()).hexdigest()[:24]


def check_password():
    if st.session_state.get("role"):
        return True

    # ---- URL токен арқылы автологин (сайт ұйықтап кетіп қайта ашылғанда парольді қайта сұрамас үшін) ----
    qp_token = st.query_params.get("auth_token", "")
    if qp_token:
        manager_pwd = st.secrets.get("MANAGER_PASSWORD", "")
        if manager_pwd and qp_token == _auth_token(manager_pwd, "manager"):
            st.session_state.role = "manager"
            st.session_state.store_access = None
            return True
        for i in range(1, 21):
            sub_pwd = st.secrets.get(f"SUBMANAGER_{i}_PASSWORD", "")
            if sub_pwd and qp_token == _auth_token(sub_pwd, f"sub{i}"):
                stores_str = st.secrets.get(f"SUBMANAGER_{i}_STORES", "")
                allowed = [int(x.strip()) for x in stores_str.split(",") if x.strip().isdigit()]
                st.session_state.role = "submanager"
                st.session_state.store_access = allowed
                return True
        stores_str = st.secrets.get("STORE_NAMES", "")
        store_names = [n.strip() for n in stores_str.split(",") if n.strip()]
        for i, name in enumerate(store_names, 1):
            owner_pwd = st.secrets.get(f"STORE_{i}_PASSWORD", "")
            if owner_pwd and qp_token == _auth_token(owner_pwd, f"owner{i}"):
                st.session_state.role = "owner"
                st.session_state.store_access = i
                return True

    st.title("🔐 Wildberries Отчёт")
    st.markdown("---")
    _, col, _ = st.columns([1, 1.5, 1])
    with col:
        st.markdown("### Вход")
        with st.form("login_form"):
            pwd = st.text_input("Пароль", type="password", placeholder="••••••••")
            submitted = st.form_submit_button("Войти →", use_container_width=True)
        if submitted:
            manager_pwd = st.secrets.get("MANAGER_PASSWORD", "")
            if manager_pwd and pwd == manager_pwd:
                st.session_state.role = "manager"
                st.session_state.store_access = None
                st.query_params["auth_token"] = _auth_token(manager_pwd, "manager")
                st.rerun()
                return True
            # SUBMANAGER — бірнеше ИП көретін қосымша менеджер
            for i in range(1, 21):
                sub_pwd = st.secrets.get(f"SUBMANAGER_{i}_PASSWORD", "")
                if sub_pwd and pwd == sub_pwd:
                    stores_str = st.secrets.get(f"SUBMANAGER_{i}_STORES", "")
                    allowed = [int(x.strip()) for x in stores_str.split(",") if x.strip().isdigit()]
                    st.session_state.role = "submanager"
                    st.session_state.store_access = allowed
                    st.query_params["auth_token"] = _auth_token(sub_pwd, f"sub{i}")
                    st.rerun()
                    return True
            stores_str = st.secrets.get("STORE_NAMES", "")
            store_names = [n.strip() for n in stores_str.split(",") if n.strip()]
            for i, name in enumerate(store_names, 1):
                owner_pwd = st.secrets.get(f"STORE_{i}_PASSWORD", "")
                if owner_pwd and pwd == owner_pwd:
                    st.session_state.role = "owner"
                    st.session_state.store_access = i
                    st.query_params["auth_token"] = _auth_token(owner_pwd, f"owner{i}")
                    st.rerun()
                    return True
            st.error("Неверный пароль!")
    return False

if not check_password():
    st.stop()

def get_stores():
    names_str = st.secrets.get("STORE_NAMES", "")
    if not names_str:
        return []
    names = [n.strip() for n in names_str.split(",") if n.strip()]
    stores = []
    for i, name in enumerate(names, 1):
        stats_key = st.secrets.get(f"STORE_{i}_STATS", "")
        analytics_key = st.secrets.get(f"STORE_{i}_ANALYTICS", "")
        finance_key = st.secrets.get(f"STORE_{i}_FINANCE", "")
        feedback_key = st.secrets.get(f"STORE_{i}_FEEDBACK", "")
        marketplace_key = st.secrets.get(f"STORE_{i}_MARKETPLACE", "")
        if stats_key:
            stores.append({
                "name": name, "idx": i,
                "stats_key": stats_key,
                "analytics_key": analytics_key,
                "finance_key": finance_key,
                "feedback_key": feedback_key,
                "marketplace_key": marketplace_key,
            })
    return stores

def days_ago_str(n):
    return (datetime.now() - timedelta(days=n)).strftime("%Y-%m-%dT00:00:00")

def wb_get_retry(url, key, params={}, max_retries=3, store_name=""):
    for attempt in range(max_retries):
        r = requests.get(url, headers={"Authorization": key}, params=params, timeout=60)
        if r.status_code == 429:
            st.info(f"⏳ [{store_name}] WB API лимит — 65 сек...")
            time.sleep(65)
            continue
        r.raise_for_status()
        return r.json()
    raise Exception("Лимит запросов превышен.")

def fetch_warehouse_remains(analytics_key):
    base = "https://seller-analytics-api.wildberries.ru"
    for attempt in range(3):
        r = requests.get(f"{base}/api/v1/warehouse_remains",
                         headers={"Authorization": analytics_key},
                         params={"groupBySa": "true"}, timeout=30)
        if r.status_code == 429:
            st.info("⏳ WB Analytics лимит — ждём 65 сек...")
            time.sleep(65)
            continue
        r.raise_for_status()
        break
    else:
        raise Exception("Превышен лимит запросов WB Analytics")

    task_id = r.json()["data"]["taskId"]
    for _ in range(30):
        time.sleep(5)
        r2 = requests.get(f"{base}/api/v1/warehouse_remains/tasks/{task_id}/status",
                          headers={"Authorization": analytics_key}, timeout=30)
        if r2.status_code == 429:
            time.sleep(65)
            continue
        r2.raise_for_status()
        status = r2.json()["data"]["status"]
        if status == "done":
            break
        elif status in ["failed", "error"]:
            raise Exception(f"Ошибка задачи: {status}")

    for attempt in range(3):
        r3 = requests.get(f"{base}/api/v1/warehouse_remains/tasks/{task_id}/download",
                          headers={"Authorization": analytics_key}, timeout=60)
        if r3.status_code == 429:
            time.sleep(65)
            continue
        r3.raise_for_status()
        return r3.json()
    raise Exception("Не удалось скачать остатки")

def parse_remains(data):
    rows = []
    for item in data:
        sa = item.get("vendorCode", "")
        in_way_to_client = 0
        in_way_return = 0
        total_stock = 0
        for wh in item.get("warehouses", []):
            name = wh.get("warehouseName", "")
            qty = wh.get("quantity", 0)
            if name == "В пути до получателей":
                in_way_to_client = qty
            elif name == "В пути возвраты на склад WB":
                in_way_return = qty
            elif name == "Всего находится на складах":
                total_stock = qty
        rows.append({
            "supplierArticle": sa,
            "qty": total_stock,
            "in_way_client": in_way_to_client + in_way_return,
            "in_way_return": in_way_return,
        })
    return pd.DataFrame(rows)

def fetch_report_detail(stats_key, date_from, date_to, store_name=""):
    urls = [
        "https://statistics-api.wildberries.ru/api/v5/supplier/reportDetailByPeriod",
        "https://statistics-api.wildberries.ru/api/v1/supplier/reportDetailByPeriod",
    ]
    all_rows = []
    for url in urls:
        try:
            rrdid = 0
            while True:
                params = {
                    "dateFrom": date_from,
                    "dateTo": date_to,
                    "rrdid": rrdid,
                    "limit": 100000
                }
                r = requests.get(url, headers={"Authorization": stats_key},
                                 params=params, timeout=60)
                if r.status_code == 404:
                    break
                if r.status_code == 429:
                    time.sleep(65)
                    continue
                r.raise_for_status()
                data = r.json()
                if not data:
                    break
                all_rows.extend(data)
                rrdid = data[-1].get("rrd_id", 0)
                if len(data) < 100000:
                    break
                time.sleep(2)
            if all_rows:
                break
        except Exception:
            continue
    return all_rows

def parse_finance(rows):
    if not rows:
        return {}
    result = {
        "for_pay": 0,        # К перечислению продавцу
        "realizacia": 0,     # Вайлдберриз реализовал Товар (Пр)
        "komissiya": 0,      # Компенсация платёжных услуг
        "vozmesh_pvz": 0,    # Возмещение за выдачу и возврат на ПВЗ
        "ads": 0,            # Удержания (реклама)
        "storage": 0,        # Хранение
        "penalty": 0,        # Штраф
        "logistic": 0,       # Услуги по доставке (доставка ВБ)
        "priemka": 0,        # Операции на приёмке
        "vozvrat": 0,        # Возврат (сумма)
        "vozvrat_qty": 0,
        "total_qty": 0,      # сатылған шт (продажа)
        "by_article": {}
    }

    for row in rows:
        oper = str(row.get("supplier_oper_name", "")).strip()
        oper_up = oper.upper()
        ppvz = float(row.get("ppvz_for_pay", 0) or 0)
        deduct = float(row.get("deduction", 0) or 0)
        storage_fee = float(row.get("storage_fee", 0) or 0)
        penalty_val = float(row.get("penalty", 0) or 0)
        delivery_rub = float(row.get("delivery_rub", 0) or 0)
        retail = float(row.get("retail_amount", 0) or 0)
        acquiring = float(row.get("acquiring_fee", 0) or 0)
        rebill = float(row.get("rebill_logistic_cost", 0) or 0)
        qty = int(row.get("quantity", 0) or 0)
        article = str(row.get("sa_name", "") or "").strip()
        nm_id = row.get("nm_id", "") or row.get("nmId", "") or ""

        if oper_up in ("ПРОДАЖА", "ДОБРОВОЛЬНАЯ КОМПЕНСАЦИЯ ПРИ ВОЗВРАТЕ"):
            result["for_pay"] += ppvz
            result["realizacia"] += retail
            result["komissiya"] += abs(acquiring)
            result["vozmesh_pvz"] += abs(rebill)
            result["total_qty"] += qty
            if article:
                if article not in result["by_article"]:
                    result["by_article"][article] = {
                        "qty": 0, "for_pay": 0, "vozvrat": 0, "vozvrat_qty": 0,
                        "realizacia": 0, "nm_id": nm_id
                    }
                result["by_article"][article]["qty"] += qty
                result["by_article"][article]["for_pay"] += ppvz
                result["by_article"][article]["realizacia"] += retail
                if nm_id and not result["by_article"][article].get("nm_id"):
                    result["by_article"][article]["nm_id"] = nm_id
        elif oper_up == "ВОЗВРАТ":
            result["for_pay"] += ppvz          # К перечислению-ге возврат сомасы да қосылады (Excel Общий итог)
            result["vozvrat"] += abs(ppvz)
            result["vozvrat_qty"] += qty
            result["komissiya"] += abs(acquiring)
            result["vozmesh_pvz"] += abs(rebill)
            if article:
                if article not in result["by_article"]:
                    result["by_article"][article] = {
                        "qty": 0, "for_pay": 0, "vozvrat": 0, "vozvrat_qty": 0,
                        "realizacia": 0, "nm_id": nm_id
                    }
                result["by_article"][article]["for_pay"] += ppvz
                result["by_article"][article]["vozvrat"] += abs(ppvz)
                result["by_article"][article]["vozvrat_qty"] += qty
        elif oper_up == "ЛОГИСТИКА" or "ДОСТАВК" in oper_up:
            result["logistic"] += abs(delivery_rub)
        elif "ОБРАБОТКА" in oper_up:
            priemka_val = abs(float(row.get("acceptance", 0) or 0)) or abs(deduct)
            result["priemka"] += priemka_val
        elif "ХРАНЕНИЕ" in oper_up:
            result["storage"] += abs(storage_fee) + abs(deduct)
        elif "ШТРАФ" in oper_up:
            result["penalty"] += abs(penalty_val) + abs(deduct)
        elif "УДЕРЖАНИЕ" in oper_up:
            result["ads"] += abs(deduct) + abs(ppvz)

    if result["logistic"] == 0:
        for row in rows:
            result["logistic"] += abs(float(row.get("delivery_rub", 0) or 0))
    return result

def status_label(q):
    if q == 0:    return "Ноль"
    if q <= 200:  return "Мало"
    if q <= 500:  return "Хорошо"
    return "Достаточно"

def load_store_data(store):
    idx = store["idx"]
    stats_key = store["stats_key"]
    analytics_key = store["analytics_key"]
    name = store["name"]
    errors = []

    if analytics_key:
        try:
            with st.spinner(f"[{name}] Загрузка остатков и FBO..."):
                remains_data = fetch_warehouse_remains(analytics_key)
                agg = parse_remains(remains_data)
        except Exception as e:
            errors.append(f"Остатки: {e}")
            agg = pd.DataFrame()
    else:
        try:
            with st.spinner(f"[{name}] Загрузка остатков..."):
                stocks_raw = wb_get_retry(
                    "https://statistics-api.wildberries.ru/api/v1/supplier/stocks",
                    stats_key, {"dateFrom": "2019-01-01"}, store_name=name
                )
                s_df = pd.DataFrame(stocks_raw)
                agg = s_df.groupby("supplierArticle").agg(
                    qty=("quantity", "sum"),
                    in_way_client=("inWayToClient", "sum"),
                ).reset_index()
                agg["in_way_return"] = 0
        except Exception as e:
            errors.append(f"Остатки: {e}")
            agg = pd.DataFrame()

    time.sleep(3)

    try:
        with st.spinner(f"[{name}] Загрузка продаж..."):
            sales_raw = wb_get_retry(
                "https://statistics-api.wildberries.ru/api/v1/supplier/sales",
                stats_key, {"dateFrom": days_ago_str(20), "flag": 0}, store_name=name
            )
            sales_df = pd.DataFrame(sales_raw) if sales_raw else pd.DataFrame()
            if not sales_df.empty and "supplierArticle" in sales_df.columns:
                if "saleID" in sales_df.columns:
                    sales_df = sales_df[~sales_df["saleID"].astype(str).str.startswith("R")]
                active = set(sales_df["supplierArticle"].unique())
                cutoff7 = datetime.now() - timedelta(days=7)
                if "date" in sales_df.columns:
                    sales_df["date"] = pd.to_datetime(sales_df["date"], errors="coerce")
                    s7 = sales_df[sales_df["date"] >= cutoff7]
                else:
                    s7 = sales_df
                daily = s7.groupby("supplierArticle").size().div(7).reset_index()
                daily.columns = ["supplierArticle", "daily_avg"]
            else:
                active = set()
                daily = pd.DataFrame(columns=["supplierArticle", "daily_avg"])
    except Exception as e:
        errors.append(f"Продажи: {e}")
        active = set()
        daily = pd.DataFrame(columns=["supplierArticle", "daily_avg"])

    time.sleep(3)

    sales30 = pd.DataFrame()
    try:
        with st.spinner(f"[{name}] Загрузка аналитики..."):
            orders_raw = wb_get_retry(
                "https://statistics-api.wildberries.ru/api/v1/supplier/orders",
                stats_key, {"dateFrom": days_ago_str(90), "flag": 0}, store_name=name
            )
            o_df = pd.DataFrame(orders_raw) if orders_raw else pd.DataFrame()
            if not o_df.empty and "date" in o_df.columns:
                o_df["date"] = pd.to_datetime(o_df["date"], errors="coerce")
                cutoff30 = datetime.now() - timedelta(days=30)
                o_df = o_df[o_df["date"] >= cutoff30]
                if "isCancel" in o_df.columns:
                    o_df = o_df[o_df["isCancel"] == False]
                o_df["date_only"] = o_df["date"].dt.date
                o_df["priceWithDisc"] = pd.to_numeric(o_df.get("priceWithDisc", 0), errors="coerce").fillna(0)
                sales30 = o_df.groupby("date_only").agg(
                    qty=("date_only", "count"),
                    revenue=("priceWithDisc", "sum")
                ).reset_index()
                sales30.columns = ["Дата", "Заказ (шт)", "Выручка (₸)"]
                sales30 = sales30.sort_values("Дата")
    except Exception as e:
        errors.append(f"Аналитика: {e}")

    df = pd.DataFrame()
    if not agg.empty:
        df = agg.merge(daily, on="supplierArticle", how="left")
        df["daily_avg"] = df["daily_avg"].fillna(0).round(1)
        if "in_way_return" not in df.columns:
            df["in_way_return"] = 0
        df["total"] = df["qty"] + df["in_way_client"]
        df["turnover"] = df.apply(
            lambda r: round((r["qty"] + r["in_way_return"]) / r["daily_avg"])
            if r["daily_avg"] > 0 else None, axis=1
        )
        df["status"] = df["qty"].apply(status_label)
        if active:
            df = df[df["supplierArticle"].isin(active)]
        df = df.reset_index(drop=True)

    return df, sales30, errors


FEEDBACK_BASE = "https://feedbacks-api.wildberries.ru"

def fetch_feedbacks(fb_key, is_answered=False, take=20):
    try:
        r = requests.get(
            f"{FEEDBACK_BASE}/api/v1/feedbacks",
            headers={"Authorization": fb_key},
            params={"isAnswered": str(is_answered).lower(), "take": take, "skip": 0, "order": "dateDesc"},
            timeout=30
        )
        if r.status_code not in (200, 201, 204):
            st.error(f"Feedbacks API ошибка: {r.status_code} — {r.text[:200]}")
            return []
        data = r.json()
        return data.get("data", {}).get("feedbacks", []) or []
    except Exception as e:
        st.error(f"Feedbacks ошибка: {e}")
        return []

def fetch_questions(fb_key, is_answered=False, take=20):
    try:
        r = requests.get(
            f"{FEEDBACK_BASE}/api/v1/questions",
            headers={"Authorization": fb_key},
            params={"isAnswered": str(is_answered).lower(), "take": take, "skip": 0, "order": "dateDesc"},
            timeout=30
        )
        if r.status_code not in (200, 201, 204):
            st.error(f"Questions API ошибка: {r.status_code} — {r.text[:200]}")
            return []
        data = r.json()
        return data.get("data", {}).get("questions", []) or data.get("questions", []) or []
    except Exception as e:
        st.error(f"Questions ошибка: {e}")
        return []

def send_feedback_reply(fb_key, feedback_id, text):
    try:
        r = requests.patch(
            f"{FEEDBACK_BASE}/api/v1/feedbacks/answer",
            headers={"Authorization": fb_key, "Content-Type": "application/json"},
            json={"id": feedback_id, "text": text},
            timeout=30
        )
        if r.status_code not in (200, 201, 204):
            st.error(f"Ошибка ответа WB {r.status_code}: {r.text[:300]}")
        return r.status_code in (200, 201, 204)
    except Exception as e:
        st.error(f"send_feedback_reply ошибка: {e}")
        return False

def send_question_reply(fb_key, question_id, text):
    try:
        r = requests.patch(
            f"{FEEDBACK_BASE}/api/v1/questions",
            headers={"Authorization": fb_key, "Content-Type": "application/json"},
            json={"id": question_id, "answer": {"text": text}, "state": "wbRu"},
            timeout=30
        )
        if r.status_code not in (200, 201, 204):
            st.error(f"Ошибка вопроса WB {r.status_code}: {r.text[:300]}")
        return r.status_code in (200, 201, 204)
    except Exception as e:
        st.error(f"send_question_reply ошибка: {e}")
        return False

def send_feedback_complaint(fb_key, feedback_id):
    try:
        r = requests.post(
            f"{FEEDBACK_BASE}/api/v1/feedbacks/report",
            headers={"Authorization": fb_key, "Content-Type": "application/json"},
            json={"id": feedback_id, "reason": "incorrect_goods_description"},
            timeout=30
        )
        return r.status_code in (200, 201, 204)
    except:
        return False


# ============================================================
# FBS МОДУЛІ — қалдық тексеру + стикер алу (Marketplace API)
# ============================================================

def fetch_fbs_warehouses(mp_key):
    r = requests.get(
        f"{MARKETPLACE_BASE}/api/v3/warehouses",
        headers={"Authorization": mp_key},
        timeout=30,
    )
    r.raise_for_status()
    return r.json()  # [{"id":..., "name":..., "cargoType":...}, ...]


def fetch_fbs_stocks(mp_key, warehouse_id, skus):
    """skus — баркодтар тізімі (list of str)."""
    r = requests.post(
        f"{MARKETPLACE_BASE}/api/v3/stocks/{warehouse_id}",
        headers={"Authorization": mp_key, "Content-Type": "application/json"},
        json={"skus": skus},
        timeout=30,
    )
    if r.status_code == 429:
        time.sleep(10)
        r = requests.post(
            f"{MARKETPLACE_BASE}/api/v3/stocks/{warehouse_id}",
            headers={"Authorization": mp_key, "Content-Type": "application/json"},
            json={"skus": skus},
            timeout=30,
        )
    r.raise_for_status()
    return r.json().get("stocks", [])  # [{"sku":..., "amount":...}, ...]


def update_fbs_stocks(mp_key, warehouse_id, stock_updates):
    """stock_updates — [{"sku": "...", "amount": 10}, ...] қалдықты жаңарту үшін."""
    r = requests.put(
        f"{MARKETPLACE_BASE}/api/v3/stocks/{warehouse_id}",
        headers={"Authorization": mp_key, "Content-Type": "application/json"},
        json={"stocks": stock_updates},
        timeout=30,
    )
    return r.status_code in (200, 204)


def fetch_new_fbs_orders(mp_key):
    r = requests.get(
        f"{MARKETPLACE_BASE}/api/v3/orders/new",
        headers={"Authorization": mp_key},
        timeout=30,
    )
    r.raise_for_status()
    return r.json().get("orders", [])


def fetch_orders_status(mp_key, order_ids):
    """Тапсырыс статусын тексеру (мысалы, жіберілді ме, отменен бе)."""
    r = requests.post(
        f"{MARKETPLACE_BASE}/api/v3/orders/status",
        headers={"Authorization": mp_key, "Content-Type": "application/json"},
        json={"orders": order_ids},
        timeout=30,
    )
    r.raise_for_status()
    return r.json().get("orders", [])


def fetch_order_stickers(mp_key, order_ids, sticker_type="png", width=58, height=40):
    """
    order_ids — сборка тапсырмаларының orderId тізімі (бір сұранымда max 100)
    sticker_type — "png" немесе "svg"
    Қайтарады: [{"orderId":..., "partA":..., "partB":..., "barcode":..., "file": "<base64>"}]
    """
    params = {"type": sticker_type, "width": width, "height": height}
    r = requests.post(
        f"{MARKETPLACE_BASE}/api/v3/orders/stickers",
        headers={"Authorization": mp_key, "Content-Type": "application/json"},
        params=params,
        json={"orders": order_ids},
        timeout=30,
    )
    if r.status_code == 429:
        time.sleep(10)
        r = requests.post(
            f"{MARKETPLACE_BASE}/api/v3/orders/stickers",
            headers={"Authorization": mp_key, "Content-Type": "application/json"},
            params=params,
            json={"orders": order_ids},
            timeout=30,
        )
    r.raise_for_status()
    return r.json().get("stickers", [])


def stickers_to_zip(stickers):
    """Барлық стикерлерді бір ZIP файлға жинайды (файл ретінде жүктеп алу үшін)."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        for s in stickers:
            order_id = s.get("orderId", "unknown")
            file_b64 = s.get("file", "")
            ext = "png"
            if file_b64:
                zf.writestr(f"sticker_{order_id}.{ext}", base64.b64decode(file_b64))
    buf.seek(0)
    return buf


def stickers_to_zip_by_cell(stickers, orders_by_id, cell_map, propusk_bytes=None, list_bytes=None):
    """
    Стикерлерді ячейка/ИП бойынша топтап ZIP жасайды.
    Қапшық аты: "{ячейка} Ячейка {ИП}", ішінде:
      "{ИП} стикер.pdf" — сол ячейкадағы барлық тапсырыстың стикерлері бір PDF-те біріктірілген
      "{ИП} пропуск.pdf" / "{ИП} Лист .pdf" — қолмен жүктелген ортақ құжаттардың көшірмесі
    orders_by_id: {orderId: order_dict} — баркодты табу үшін
    cell_map: {barcode: {"cell": "10", "ip": "Тендернес"}}
    Қайтарады: (BytesIO буфер, сәйкессіз қалған стикерлер саны)
    """
    grouped = {}
    unmatched = 0
    for s in stickers:
        order_id = s.get("orderId")
        file_b64 = s.get("file", "")
        if not file_b64:
            continue
        order = orders_by_id.get(order_id, {})
        skus = order.get("skus", [])
        barcode = skus[0] if skus else ""
        cell_info = cell_map.get(str(barcode))
        if cell_info:
            cell = cell_info.get("cell", "")
            ip = cell_info.get("ip", "")
        else:
            cell, ip = "", "Белгісіз"
            unmatched += 1
        grouped.setdefault((cell, ip), []).append(base64.b64decode(file_b64))

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        for (cell, ip), img_bytes_list in grouped.items():
            folder = f"{cell} Ячейка {ip}".strip()

            images = [Image.open(io.BytesIO(b)).convert("RGB") for b in img_bytes_list]
            if images:
                sticker_pdf_buf = io.BytesIO()
                images[0].save(sticker_pdf_buf, format="PDF", save_all=True, append_images=images[1:])
                zf.writestr(f"{folder}/{ip} стикер.pdf", sticker_pdf_buf.getvalue())

            if propusk_bytes:
                zf.writestr(f"{folder}/{ip} пропуск.pdf", propusk_bytes)
            if list_bytes:
                zf.writestr(f"{folder}/{ip} Лист .pdf", list_bytes)

    buf.seek(0)
    return buf, unmatched


def show_cell_map_editor(store):
    """Баркод → Ячейка → ИП картасын қолмен енгізу кестесі."""
    idx = store["idx"]
    all_cells = load_json(CELL_MAP_FILE)
    store_cells = all_cells.get(str(idx), {})

    rows = [{"Баркод": bc, "Ячейка": v.get("cell", ""), "ИП": v.get("ip", "")} for bc, v in store_cells.items()]
    if not rows:
        rows = [{"Баркод": "", "Ячейка": "", "ИП": ""}]
    cell_df = pd.DataFrame(rows)

    st.caption("Google Sheets-тегі кестеңізден Баркод, Ячейка, ИП бағандарын осында көшіріп қоюға болады")
    edited_cells = st.data_editor(
        cell_df, use_container_width=True, height=300, num_rows="dynamic", key=f"cell_map_editor_{idx}",
        column_config={
            "Баркод": st.column_config.TextColumn(),
            "Ячейка": st.column_config.TextColumn(),
            "ИП": st.column_config.TextColumn(),
        },
    )
    new_cells = {}
    for _, rr in edited_cells.iterrows():
        bc = str(rr["Баркод"]).strip()
        if bc and bc != "nan":
            new_cells[bc] = {"cell": str(rr["Ячейка"]).strip(), "ip": str(rr["ИП"]).strip()}
    if new_cells != store_cells:
        all_cells[str(idx)] = new_cells
        save_json(CELL_MAP_FILE, all_cells)
    return new_cells


EXPIRY_FILE = os.path.join(DATA_DIR, "expiry_data.json")
DELIVERED_SUPPLIES_FILE = os.path.join(DATA_DIR, "delivered_supplies.json")
CELL_MAP_FILE = os.path.join(DATA_DIR, "cell_map.json")
PERSIST_FILES = PERSIST_FILES + ("expiry_data.json", "delivered_supplies.json", "cell_map.json")


def fetch_all_orders(mp_key, limit=1000, next_id=0):
    r = requests.get(
        f"{MARKETPLACE_BASE}/api/v3/orders",
        headers={"Authorization": mp_key},
        params={"limit": limit, "next": next_id},
        timeout=30,
    )
    r.raise_for_status()
    return r.json()  # {"orders": [...], "next": ...}


def create_supply(mp_key, name):
    r = requests.post(
        f"{MARKETPLACE_BASE}/api/v3/supplies",
        headers={"Authorization": mp_key, "Content-Type": "application/json"},
        json={"name": name},
        timeout=30,
    )
    r.raise_for_status()
    return r.json().get("id")


def get_supplies(mp_key, limit=200):
    r = requests.get(
        f"{MARKETPLACE_BASE}/api/v3/supplies",
        headers={"Authorization": mp_key},
        params={"limit": limit, "next": 0},
        timeout=30,
    )
    r.raise_for_status()
    return r.json().get("supplies", [])


def add_orders_to_supply(mp_key, supply_id, order_ids):
    """Бір поставкаға бірнеше сборочное задание қосады (максимум 100 бір сұранымда)."""
    ok_all = True
    for i in range(0, len(order_ids), 100):
        batch = order_ids[i:i + 100]
        r = requests.patch(
            f"{MARKETPLACE_BASE}/api/v3/supplies/{supply_id}/orders",
            headers={"Authorization": mp_key, "Content-Type": "application/json"},
            json={"orders": batch},
            timeout=30,
        )
        if r.status_code not in (200, 204):
            ok_all = False
    return ok_all


def deliver_supply(mp_key, supply_id):
    r = requests.patch(
        f"{MARKETPLACE_BASE}/api/v3/supplies/{supply_id}/deliver",
        headers={"Authorization": mp_key},
        timeout=30,
    )
    return r.status_code in (200, 204)



    """Биту күніне дейін қанша күн қалғанын және статус түсін қайтарады."""
    if not exp_date_str:
        return None, ""
    try:
        exp_date = datetime.strptime(exp_date_str, "%Y-%m-%d").date()
    except Exception:
        return None, ""
    days_left = (exp_date - date.today()).days
    if days_left < 0:
        return days_left, "background-color:#FCEBEB;color:#A32D2D;font-weight:bold"
    elif days_left <= 30:
        return days_left, "background-color:#FCEBEB;color:#A32D2D;font-weight:bold"
    elif days_left <= 90:
        return days_left, "background-color:#FAEEDA;color:#854F0B;font-weight:bold"
    else:
        return days_left, "background-color:#EAF3DE;color:#3B6D11"


def show_fbs_tab(store):
    idx = store["idx"]
    mp_key = store.get("marketplace_key", "")

    if not mp_key:
        st.warning(f"⚠️ Бұл кабинетте Marketplace токені жоқ. Secrets-ке STORE_{idx}_MARKETPLACE қосыңыз.")
        st.caption("Токенді ЛК Wildberries → Настройки → Доступ к новому API бөлімінен, Marketplace категориясын таңдап жасаңыз.")
        return

    sub1, sub2 = st.tabs(["📥 Тапсырыстар & Стикер", "📦 Қалдықты тексеру"])

    with st.expander("🗂️ Ячейка картасы (Баркод → Ячейка → ИП)", expanded=False):
        cell_map_current = show_cell_map_editor(store)

    # ---------- ТАПСЫРЫСТАР (Новые / На сборке / В доставке) + СТИКЕР ----------
    with sub1:
        if st.button("🔄 Тапсырыстарды жаңалау", key=f"fbs_orders_load_{idx}"):
            with st.spinner("Жүктелуде..."):
                try:
                    new_orders = fetch_new_fbs_orders(mp_key)
                    all_data = fetch_all_orders(mp_key, limit=1000)
                    all_orders = all_data.get("orders", [])
                    new_ids = [o.get("id") for o in new_orders]
                    non_new_orders = [o for o in all_orders if o.get("id") not in new_ids]
                    order_ids_check = [o.get("id") for o in non_new_orders]
                    status_map = {}
                    for i in range(0, len(order_ids_check), 1000):
                        batch = order_ids_check[i:i + 1000]
                        if batch:
                            statuses = fetch_orders_status(mp_key, batch)
                            for s in statuses:
                                status_map[s.get("id")] = s
                    st.session_state[f"fbs_new_{idx}"] = new_orders
                    st.session_state[f"fbs_nonnew_{idx}"] = non_new_orders
                    st.session_state[f"fbs_status_map_{idx}"] = status_map
                except Exception as e:
                    st.error(f"Қате: {e}")

        new_orders = st.session_state.get(f"fbs_new_{idx}", [])
        non_new_orders = st.session_state.get(f"fbs_nonnew_{idx}", [])
        status_map = st.session_state.get(f"fbs_status_map_{idx}", {})

        all_delivered = load_json(DELIVERED_SUPPLIES_FILE)
        delivered_supplies = set(all_delivered.get(str(idx), []))

        confirm_orders = []
        deliver_orders = []
        for o in non_new_orders:
            st_info = status_map.get(o.get("id"), {})
            sup_status = st_info.get("supplierStatus", "")
            supply_id = o.get("supplyId", "")
            if sup_status == "confirm":
                if supply_id and supply_id in delivered_supplies:
                    deliver_orders.append(o)
                else:
                    confirm_orders.append(o)

        n_new, n_confirm, n_deliver = len(new_orders), len(confirm_orders), len(deliver_orders)

        choice = st.radio(
            "Статус",
            [f"🆕 Новые ({n_new})", f"📦 На сборке ({n_confirm})", f"🚚 В доставке ({n_deliver})"],
            horizontal=True, key=f"fbs_status_choice_{idx}", label_visibility="collapsed",
        )

        all_expiry = load_json(EXPIRY_FILE)
        store_expiry = all_expiry.get(str(idx), {})

        def build_rows(orders_list):
            rows = []
            for o in orders_list:
                order_id = str(o.get("id"))
                saved_exp = store_expiry.get(order_id, "")
                exp_val = None
                if saved_exp:
                    try:
                        exp_val = datetime.strptime(saved_exp, "%Y-%m-%d").date()
                    except Exception:
                        exp_val = None
                rows.append({
                    "Таңдау": False,
                    "orderId": o.get("id"),
                    "Артикул": o.get("article", ""),
                    "Баркод": o.get("skus", [""])[0] if o.get("skus") else "",
                    "Құн (₸)": o.get("convertedPrice", 0) / 100 if o.get("convertedPrice") else 0,
                    "Құрылған күні": o.get("createdAt", "")[:16].replace("T", " "),
                    "Сроку годности": exp_val,
                })
            return rows

        # ================= НОВЫЕ =================
        if choice.startswith("🆕"):
            if not new_orders:
                st.info("Жаңа тапсырыс жоқ. 👆 «Тапсырыстарды жаңалау» басыңыз")
            else:
                rows = build_rows(new_orders)

                unique_articles = sorted(set(r["Артикул"] for r in rows if r["Артикул"]))
                bc1, bc2, bc3 = st.columns([1.3, 1, 0.7])
                with bc1:
                    bulk_article = st.selectbox("Артикул таңдаңыз", unique_articles, key=f"fbs_bulk_art_{idx}")
                with bc2:
                    bulk_date = st.date_input("Сроку годности", value=date.today()+timedelta(days=180), key=f"fbs_bulk_date_{idx}")
                with bc3:
                    st.markdown("<br>", unsafe_allow_html=True)
                    bulk_apply = st.button("✅ Барлығына қолдану", key=f"fbs_bulk_apply_{idx}", use_container_width=True)

                if bulk_apply and bulk_article:
                    count_applied = 0
                    for r in rows:
                        if r["Артикул"] == bulk_article:
                            store_expiry[str(r["orderId"])] = bulk_date.strftime("%Y-%m-%d")
                            count_applied += 1
                    all_expiry[str(idx)] = store_expiry
                    save_json(EXPIRY_FILE, all_expiry)
                    st.success(f"✅ «{bulk_article}» — {count_applied} тапсырысқа {bulk_date.strftime('%d.%m.%Y')} қойылды")
                    st.rerun()

                st.divider()
                for row in rows:
                    saved_exp2 = store_expiry.get(str(row["orderId"]), "")
                    if saved_exp2:
                        try:
                            row["Сроку годности"] = datetime.strptime(saved_exp2, "%Y-%m-%d").date()
                        except Exception:
                            pass
                df_orders = pd.DataFrame(rows)

                edited = st.data_editor(
                    df_orders, use_container_width=True, height=400, key=f"fbs_new_editor_{idx}",
                    column_config={
                        "Таңдау": st.column_config.CheckboxColumn(),
                        "orderId": st.column_config.NumberColumn(disabled=True),
                        "Артикул": st.column_config.TextColumn(disabled=True),
                        "Баркод": st.column_config.TextColumn(disabled=True),
                        "Құн (₸)": st.column_config.NumberColumn(disabled=True, format="%.0f ₸"),
                        "Құрылған күні": st.column_config.TextColumn(disabled=True),
                        "Сроку годности": st.column_config.DateColumn(format="DD.MM.YYYY"),
                    },
                )

                new_expiry = {}
                for _, rr in edited.iterrows():
                    if pd.notna(rr["Сроку годности"]):
                        new_expiry[str(rr["orderId"])] = rr["Сроку годности"].strftime("%Y-%m-%d")
                for k, v in new_expiry.items():
                    store_expiry[k] = v
                if new_expiry:
                    all_expiry[str(idx)] = store_expiry
                    save_json(EXPIRY_FILE, all_expiry)

                selected_ids = edited[edited["Таңдау"]]["orderId"].tolist()
                st.caption(f"Таңдалды: {len(selected_ids)} тапсырыс")

                c1, c2 = st.columns(2)
                with c1:
                    if st.button("📦 Сборкаға қосу (таңдалғандарды)", key=f"fbs_to_confirm_{idx}", disabled=not selected_ids, use_container_width=True):
                        with st.spinner("Поставка жасалуда..."):
                            try:
                                supply_name = f"{store['name']} — {datetime.now().strftime('%d.%m.%Y %H:%M')}"
                                supply_id = create_supply(mp_key, supply_name)
                                ok = add_orders_to_supply(mp_key, supply_id, [int(x) for x in selected_ids])
                                if ok:
                                    st.success(f"✅ {len(selected_ids)} тапсырыс жаңа поставкаға ({supply_id}) қосылды — олар «На сборке» бөліміне өтті")
                                else:
                                    st.warning("⚠️ Кейбір тапсырыстар қосылмады, тексеріңіз")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Қате: {e}")
                with c2:
                    if st.button("🖨️ Стикер алу (таңдалғандарға)", key=f"fbs_sticker_btn_{idx}", disabled=not selected_ids, use_container_width=True):
                        with st.spinner("Стикерлер алынуда..."):
                            try:
                                stickers = fetch_order_stickers(mp_key, selected_ids, sticker_type="png")
                                st.session_state[f"fbs_stickers_{idx}"] = stickers
                                st.success(f"✅ {len(stickers)} стикер алынды!")
                            except Exception as e:
                                st.error(f"Стикер алу қатесі: {e}")

                stickers = st.session_state.get(f"fbs_stickers_{idx}", [])
                if stickers:
                    zip_buf = stickers_to_zip(stickers)
                    st.download_button(
                        "⬇️ Барлық стикерді ZIP түрінде жүктеу (жай)",
                        data=zip_buf.getvalue(),
                        file_name=f"stickers_{store['name']}.zip",
                        mime="application/zip",
                        key=f"fbs_zip_dl_{idx}",
                    )

                    st.divider()
                    st.markdown("**🗂️ Ячейка/ИП бойынша топталған ZIP (стикер + пропуск + лист подбора)**")
                    pc1, pc2 = st.columns(2)
                    with pc1:
                        propusk_file = st.file_uploader("Пропуск PDF (поставкаға, міндетті емес)", type=["pdf"], key=f"fbs_propusk_{idx}")
                    with pc2:
                        list_file = st.file_uploader("Лист подбора PDF (поставкаға, міндетті емес)", type=["pdf"], key=f"fbs_list_{idx}")

                    if st.button("📦 Ячейка бойынша ZIP жасау", key=f"fbs_zip_cell_{idx}", use_container_width=True):
                        orders_by_id = {o.get("id"): o for o in new_orders}
                        propusk_bytes = propusk_file.read() if propusk_file else None
                        list_bytes = list_file.read() if list_file else None
                        zip_cell_buf, unmatched = stickers_to_zip_by_cell(
                            stickers, orders_by_id, cell_map_current, propusk_bytes, list_bytes
                        )
                        st.session_state[f"fbs_zip_cell_buf_{idx}"] = zip_cell_buf.getvalue()
                        if unmatched:
                            st.warning(f"⚠️ {unmatched} стикердің баркоды ячейка картасында табылмады — «belgisiz_yacheyka» қапшығына салынды")
                        else:
                            st.success("✅ Барлық стикер ячейкасы бойынша топталды")

                    zip_cell_data = st.session_state.get(f"fbs_zip_cell_buf_{idx}")
                    if zip_cell_data:
                        st.download_button(
                            "⬇️ Ячейка бойынша ZIP жүктеу",
                            data=zip_cell_data,
                            file_name=f"yacheyka_{store['name']}.zip",
                            mime="application/zip",
                            key=f"fbs_zip_cell_dl_{idx}",
                        )

                    cols = st.columns(4)
                    for i, s in enumerate(stickers):
                        with cols[i % 4]:
                            if s.get("file"):
                                st.image(base64.b64decode(s["file"]), caption=f"orderId: {s.get('orderId')}")

        # ================= НА СБОРКЕ =================
        elif choice.startswith("📦"):
            if not confirm_orders:
                st.info("«На сборке» статусында тапсырыс жоқ")
            else:
                by_supply = {}
                for o in confirm_orders:
                    sid = o.get("supplyId", "—")
                    by_supply.setdefault(sid, []).append(o)

                for sid, orders_in_supply in by_supply.items():
                    with st.container():
                        st.markdown(f"**Поставка: `{sid}`** — {len(orders_in_supply)} тапсырыс")
                        rows = build_rows(orders_in_supply)
                        for row in rows:
                            saved_exp2 = store_expiry.get(str(row["orderId"]), "")
                            if saved_exp2:
                                try:
                                    row["Сроку годности"] = datetime.strptime(saved_exp2, "%Y-%m-%d").date()
                                except Exception:
                                    pass
                        st.dataframe(pd.DataFrame(rows).drop(columns=["Таңдау"]), use_container_width=True, height=min(300, 45+len(rows)*35))

                        if sid != "—" and st.button("🚚 Осы поставканы доставкаға беру", key=f"fbs_deliver_{idx}_{sid}"):
                            with st.spinner("Доставкаға берілуде..."):
                                ok = deliver_supply(mp_key, sid)
                                if ok:
                                    delivered_supplies.add(sid)
                                    all_delivered[str(idx)] = list(delivered_supplies)
                                    save_json(DELIVERED_SUPPLIES_FILE, all_delivered)
                                    st.success(f"✅ Поставка {sid} доставкаға берілді")
                                    st.rerun()
                                else:
                                    st.error("Қате — поставканы доставкаға беру мүмкін болмады")
                        st.divider()

        # ================= В ДОСТАВКЕ =================
        else:
            if not deliver_orders:
                st.info("«В доставке» статусында тапсырыс жоқ")
            else:
                rows = build_rows(deliver_orders)
                for row in rows:
                    saved_exp2 = store_expiry.get(str(row["orderId"]), "")
                    if saved_exp2:
                        try:
                            row["Сроку годности"] = datetime.strptime(saved_exp2, "%Y-%m-%d").date()
                        except Exception:
                            pass
                st.dataframe(pd.DataFrame(rows).drop(columns=["Таңдау"]), use_container_width=True, height=min(400, 45+len(rows)*35))

    # ---------- ҚАЛДЫҚТЫ ТЕКСЕРУ ----------
    with sub2:
        try:
            warehouses = fetch_fbs_warehouses(mp_key)
        except Exception as e:
            st.error(f"Складтарды алу қатесі: {e}")
            warehouses = []

        if not warehouses:
            st.info("Складтар табылмады")
        else:
            wh_names = [f"{w['name']} (ID: {w['id']})" for w in warehouses]
            sel_wh = st.selectbox("Склад таңдаңыз", wh_names, key=f"fbs_wh_sel_{idx}")
            wh_id = warehouses[wh_names.index(sel_wh)]["id"]

            skus_input = st.text_area(
                "Баркодтарды енгізіңіз (әр жолда біреу)",
                key=f"fbs_skus_input_{idx}",
                height=120,
                placeholder="2037654321987\n2037654321988",
            )

            if st.button("🔍 Қалдықты тексеру", key=f"fbs_stock_check_{idx}"):
                skus = [s.strip() for s in skus_input.split("\n") if s.strip()]
                if not skus:
                    st.warning("Ең болмаса бір баркод енгізіңіз")
                else:
                    with st.spinner("Тексерілуде..."):
                        try:
                            stocks = fetch_fbs_stocks(mp_key, wh_id, skus)
                            if stocks:
                                st.dataframe(pd.DataFrame(stocks), use_container_width=True)
                            else:
                                st.warning("Бұл баркодтар бойынша қалдық табылмады")
                        except Exception as e:
                            st.error(f"Қате: {e}")


def ai_generate_reply(product_name, review_text, rating, reply_type="feedback", pros="", cons="", bables="", order_status=""):
    try:
        if reply_type == "feedback":
            is_rejected = order_status in ("rejected", "cancelled", "canceled")
            system = (
                "Ты — вежливый менеджер магазина косметики на Wildberries.\n"
                "Правила:\n"
                "- Отвечай строго по содержанию отзыва\n"
                "- 2-3 предложения максимум\n"
                "- Не копируй текст отзыва обратно\n"
                "- Не используй: Спасибо за отзыв, Мы рады, Будем рады видеть вас снова\n"
                + ("- Покупатель ОТКАЗАЛСЯ от товара или сделал ВОЗВРАТ. Не предлагай замену или другой заказ. Извинись и предложи обратиться в поддержку WB.\n" if is_rejected else
                   "- Если отзыв положительный — коротко подтверди и поблагодари\n"
                   "- Если негативный — извинись, предложи решение (замену, возврат, связаться с поддержкой)\n"
                   "- Если смешанный — отреагируй на минус и похвали плюс\n"
                   "- НЕ рекомендуй другие товары и продукты ни при каких условиях\n")
                + "- Пиши только на русском языке"
            )
            parts = ["Товар: " + product_name, "Оценка: " + str(rating) + " из 5"]
            if order_status in ("rejected", "cancelled", "canceled"):
                parts.append("Статус заказа: ОТКАЗ/ВОЗВРАТ")
            if pros:
                parts.append("Плюсы: " + pros)
            if cons:
                parts.append("Минусы: " + cons)
            if bables:
                parts.append("Жалобы покупателя: " + bables)
            if review_text:
                parts.append("Комментарий: " + review_text)
            parts.append("\nНапиши ответ продавца:")
            prompt = "\n".join(parts)
        else:
            system = (
                "Ты — компетентный менеджер магазина косметики на Wildberries.\n"
                "Правила:\n"
                "- Отвечай КОНКРЕТНО на заданный вопрос\n"
                "- Если вопрос про конкретный товар — дай полезный ответ про этот товар\n"
                "- Если не знаешь точный ответ — направь к описанию товара или предложи связаться\n"
                "- 2-3 предложения максимум\n"
                "- Не используй шаблонные фразы\n"
                "- Пиши только на русском языке"
            )
            prompt = "Товар: " + product_name + "\nВопрос покупателя: " + review_text + "\n\nНапиши ответ продавца на вопрос:"

        anthropic_key = st.secrets.get("ANTHROPIC_API_KEY", "")
        if not anthropic_key:
            return "⚠️ ANTHROPIC_API_KEY Secrets-ке қосылмаған"

        for attempt in range(3):
            r = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "Content-Type": "application/json",
                    "x-api-key": anthropic_key,
                    "anthropic-version": "2023-06-01"
                },
                json={
                    "model": "claude-sonnet-4-5",
                    "max_tokens": 300,
                    "system": system,
                    "messages": [{"role": "user", "content": prompt}]
                },
                timeout=30
            )
            if r.status_code == 529:
                time.sleep(10)
                continue
            r.raise_for_status()
            return r.json()["content"][0]["text"].strip()
        return "⚠️ Anthropic сервері бос емес, кейінірек қайталаңыз"

    except Exception as e:
        return f"ИИ ошибка: {e}"

def render_stars(rating):
    filled = "★" * rating
    empty = "☆" * (5 - rating)
    color = "#E24B4A" if rating <= 3 else "#F0C040"
    return f'<span style="color:{color};font-size:15px;">{filled}{empty}</span>'

def show_sales_returns_tab(store):
    idx = store["idx"]
    stats_key = store["stats_key"]
    finance_key = store["finance_key"]
    name = store["name"]

    st.markdown("#### 🛒 Ежедневный отчёт — продажи и возвраты")
    col1, col2 = st.columns([1, 1])
    with col1:
        sel_date = st.date_input("Күн / Дата", value=date.today()-timedelta(days=1), key=f"sr_date_{idx}")
    with col2:
        st.markdown("<br>", unsafe_allow_html=True)
        load_sr = st.button("🔄 Загрузить день", key=f"sr_load_{idx}", use_container_width=True)

    sr_key = f"sales_day_{idx}"
    day_key = f"sr_day_{idx}"
    cur_day = str(sel_date)
    use_key = finance_key if finance_key else stats_key
    if st.session_state.get(day_key) != cur_day:
        if sr_key in st.session_state:
            del st.session_state[sr_key]
        st.session_state[day_key] = cur_day

    if load_sr:
        with st.spinner(f"[{name}] {sel_date} күнгі есеп..."):
            try:
                target = sel_date.strftime("%Y-%m-%d")
                d_from = (sel_date - timedelta(days=3)).strftime("%Y-%m-%d")
                d_to = (sel_date + timedelta(days=5)).strftime("%Y-%m-%d")
                wide = fetch_report_detail(use_key, d_from, d_to, store_name=name)

                rows = [r for r in (wide or []) if str(r.get("sale_dt", "") or "")[:10] == target]
                st.session_state[sr_key] = rows or []
                if rows:
                    st.success(f"✅ Загружено! ({len(rows)} операций)")
                else:
                    st.warning(f"⚠️ {target} күні деректер табылмады.")
                    if wide:
                        first = wide[0]
                        date_fields = {k: v for k, v in first.items() if "dt" in k.lower() or "date" in k.lower()}
                        st.caption(f"🔍 {len(wide)} операция бар. Күн өрістері: {date_fields}")
                        sale_dts = sorted(set(str(r.get("sale_dt","") or "")[:10] for r in wide if r.get("sale_dt")))
                        st.caption(f"🔍 Барлық sale_dt: {sale_dts}")
                    else:
                        st.caption(f"🔍 {d_from}→{d_to} аралығында reportDetailByPeriod 0 операция қайтарды (расчёт әлі жоқ).")
            except Exception as e:
                st.error(f"Ошибка: {e}")

    raw = st.session_state.get(sr_key)
    if raw is None:
        st.info("👆 Күнді таңдап **«Загрузить день»** басыңыз")
        st.caption("Финанс API — Продажа (реализация) сомасы, тиынына дейін")
        return

    if not raw:
        st.warning("Бұл күні деректер жоқ")
        return

    agg = {}
    for row in raw:
        oper = str(row.get("supplier_oper_name", "")).strip().upper()
        sa = str(row.get("sa_name", "") or "").strip()
        nm = row.get("nm_id", "") or row.get("nmId", "") or ""
        qty = int(row.get("quantity", 0) or 0)
        retail = float(row.get("retail_amount", 0) or 0)
        if not sa:
            continue
        if sa not in agg:
            agg[sa] = {"sa": sa, "nm": nm, "sold": 0, "vozvrat": 0, "revenue": 0.0, "vozvrat_sum": 0.0}
        if not agg[sa]["nm"]:
            agg[sa]["nm"] = nm
        if oper in ("ПРОДАЖА", "ДОБРОВОЛЬНАЯ КОМПЕНСАЦИЯ ПРИ ВОЗВРАТЕ"):
            agg[sa]["sold"] += qty
            agg[sa]["revenue"] += retail
        elif oper == "ВОЗВРАТ":
            agg[sa]["vozvrat"] += qty
            agg[sa]["vozvrat_sum"] += abs(retail)

    sr_rows = []
    for sa, d in agg.items():
        if d["sold"] == 0 and d["vozvrat"] == 0:
            continue
        sr_rows.append({
            "Артикул продавца": sa,
            "Артикул WB": str(d["nm"] or ""),
            "Продажа (₸)": round(d["revenue"], 2),
            "Продано (шт)": d["sold"],
            "Возврат (шт)": d["vozvrat"],
            "Итого (шт)": d["sold"] - d["vozvrat"],
        })
    if not sr_rows:
        st.warning("Бұл күні сатылым жоқ")
        return

    sr_df = pd.DataFrame(sr_rows).sort_values("Продано (шт)", ascending=False).reset_index(drop=True)
    total_sold = sr_df["Продано (шт)"].sum()
    total_vq = sr_df["Возврат (шт)"].sum()
    total_rev = sr_df["Продажа (₸)"].sum()

    c1, c2, c3 = st.columns(3)
    c1.metric("📦 Продано", f"{int(total_sold):,} шт".replace(",", " "))
    c2.metric("↩️ Возврат", f"{int(total_vq):,} шт".replace(",", " "))
    c3.metric("💰 Продажа", f"{total_rev:,.2f} ₸".replace(",", " "))
    st.caption(f"📅 {sel_date.strftime('%d.%m.%Y')} күнгі есеп")
    st.divider()

    def style_vozvrat(val):
        if pd.isna(val) or val == 0:
            return ""
        return "color:#A32D2D;font-weight:bold"
    styled_sr = sr_df.style.map(style_vozvrat, subset=["Возврат (шт)"])
    st.dataframe(styled_sr, use_container_width=True, height=460,
        column_config={"Продажа (₸)": st.column_config.NumberColumn(format="%.2f ₸")})

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        pd.DataFrame().to_excel(writer, sheet_name="_tmp", index=False)
        sr_df.to_excel(writer, index=False, sheet_name="Ежедневный отчет")
    st.download_button(f"⬇️ Excel — {sel_date.strftime('%d.%m.%Y')}", data=buf.getvalue(),
        file_name=f"Ежедневный_{store['name']}_{sel_date.strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"sr_dl_{idx}")

    vozvrat_only = sr_df[sr_df["Возврат (шт)"] > 0]
    if not vozvrat_only.empty:
        st.divider()
        st.markdown("**↩️ Возврат болған тауарлар:**")
        for _, rr in vozvrat_only.iterrows():
            st.markdown(f"- **{rr['Артикул продавца']}** (WB: {rr['Артикул WB']}) — "
                        f":red[{rr['Возврат (шт)']} шт]")

def show_feedback_tab(store):
    idx = store["idx"]
    fb_key = store.get("feedback_key", "")

    if not fb_key:
        st.warning("⚠️ Добавьте токен STORE_{n}_FEEDBACK в Secrets")
        return

    load_key = f"fb_data_{idx}"
    if load_key not in st.session_state:
        st.session_state[load_key] = None

    if st.button("🔄 Загрузить", key=f"fb_load_{idx}", use_container_width=False):
        with st.spinner("Загрузка..."):
            feedbacks = fetch_feedbacks(fb_key, is_answered=False, take=30)
            questions = fetch_questions(fb_key, is_answered=False, take=30)
            st.session_state[load_key] = {"feedbacks": feedbacks, "questions": questions}
            st.rerun()

    data = st.session_state[load_key]
    if data is None:
        st.info("👆 Нажмите кнопку **«Загрузить»**")
        return

    feedbacks = data.get("feedbacks", [])
    questions = data.get("questions", [])
    auto_replied = load_json(_tmp(f"auto_replied_{idx}.json"))

    low_star = [f for f in feedbacks if f.get("productValuation", 5) <= 3]
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("⭐ Новые отзывы", len(feedbacks))
    c2.metric("❓ Вопросы", len(questions), help="Ответьте вручную")
    c3.metric("🤖 Ответов отправлено", len(auto_replied))
    c4.metric("🔴 1-3 звезды", len(low_star))
    st.divider()

    t1, t2 = st.tabs([f"⭐ Отзывы ({len(feedbacks)})", f"❓ Авто вопросы ({len(questions)})"])

    with t1:
        if not feedbacks:
            st.success("✅ Нет неотвеченных отзывов!")
        for fb in feedbacks:
            fb_id = fb.get("id", "")
            rating = fb.get("productValuation") or fb.get("rating") or 0
            pd_ = fb.get("productDetails", {}) or {}
            text = fb.get("text", "") or ""
            bables = fb.get("bables", []) or []
            bables_text = ", ".join(bables) if bables else ""
            pros = fb.get("pros", "") or ""
            cons = fb.get("cons", "") or ""
            order_status = fb.get("orderStatus", "") or ""
            product = pd_.get("productName", "") or fb.get("productName", "") or ""
            created = fb.get("createdDate", "")[:10] if fb.get("createdDate") else ""
            preview_key_fb = f"preview_fb_{fb_id}"

            with st.container():
                col1, col2 = st.columns([6, 2])
                with col1:
                    st.markdown(render_stars(rating) + f' &nbsp; <span style="font-size:12px;color:gray;">{product} · {created}</span>', unsafe_allow_html=True)
                    if bables_text:
                        st.caption(f"⚠️ {bables_text}")
                    if text:
                        st.caption(f'"{text[:200]}{"..." if len(text)>200 else ""}"')
                with col2:
                    if rating <= 3:
                        st.caption("🔴 1-3 звезды")

                if fb_id in auto_replied:
                    reply_text = auto_replied[fb_id]
                    if "ИИ ошибка" in reply_text or "Error" in reply_text or "401" in reply_text or "404" in reply_text:
                        del auto_replied[fb_id]
                        save_json(_tmp(f"auto_replied_{idx}.json"), auto_replied)
                        st.warning("⚠️ Ошибочный ответ удалён")
                    else:
                        st.success("✅ Опубликован")
                        with st.expander("Посмотреть ответ"):
                            st.caption(reply_text)
                elif preview_key_fb in st.session_state:
                    preview_text = st.session_state[preview_key_fb]
                    edited = st.text_area("✏️ Ответ ИИ — можно редактировать:", value=preview_text, key=f"edit_fb_{fb_id}", height=100)
                    c1, c2 = st.columns(2)
                    with c1:
                        if st.button("📤 Опубликовать", key=f"pub_fb_{fb_id}", use_container_width=True):
                            with st.spinner("Отправляется..."):
                                time.sleep(1.1)
                                ok = send_feedback_reply(fb_key, fb_id, edited)
                                if ok:
                                    auto_replied[fb_id] = edited
                                    save_json(_tmp(f"auto_replied_{idx}.json"), auto_replied)
                                    del st.session_state[preview_key_fb]
                                    st.rerun()
                                else:
                                    st.error("Не отправлено")
                    with c2:
                        if st.button("🗑 Удалить", key=f"del_fb_{fb_id}", use_container_width=True):
                            del st.session_state[preview_key_fb]
                            st.rerun()
                else:
                    if st.button("🤖 Сгенерировать ответ ИИ", key=f"ai_fb_{fb_id}", use_container_width=True):
                        with st.spinner("ИИ генерирует ответ..."):
                            reply = ai_generate_reply(product, text, rating, "feedback", pros, cons, bables_text, order_status)
                            st.session_state[preview_key_fb] = reply
                            st.rerun()
                st.divider()

    with t2:
        if not questions:
            st.success("✅ Нет неотвеченных вопросов!")
        else:
            unanswered_q = [q for q in questions if q.get("id", "") not in auto_replied]
            if unanswered_q:
                st.info(f"📋 Жауапсыз вопрос: **{len(unanswered_q)}** дана")
                if st.button(f"🤖 Авто ответ барлығына ({len(unanswered_q)} вопрос)", key=f"auto_all_q_{idx}", use_container_width=True, type="primary"):
                    progress_bar = st.progress(0, text="Жіберілуде...")
                    success_count = 0
                    error_count = 0
                    for i, q in enumerate(unanswered_q):
                        q_id = q.get("id", "")
                        q_text = q.get("text", "") or ""
                        pd_q = q.get("productDetails", {}) or {}
                        product = pd_q.get("productName", "") or q.get("productName", "") or ""
                        pct = int((i) / len(unanswered_q) * 100)
                        progress_bar.progress(pct, text=f"ИИ жауап жасауда: {i+1}/{len(unanswered_q)} — {product[:40]}")
                        reply = ai_generate_reply(product, q_text, 5, "question")
                        if reply and "ошибка" not in reply.lower() and "⚠️" not in reply:
                            ok = send_question_reply(fb_key, q_id, reply)
                            if ok:
                                auto_replied[q_id] = reply
                                success_count += 1
                            else:
                                error_count += 1
                        else:
                            error_count += 1
                        time.sleep(1.5)
                    save_json(_tmp(f"auto_replied_{idx}.json"), auto_replied)
                    progress_bar.progress(100, text="Дайын!")
                    time.sleep(0.5)
                    if success_count:
                        st.success(f"✅ {success_count} вопросқа жауап жіберілді!")
                    if error_count:
                        st.warning(f"⚠️ {error_count} вопрос жіберілмеді")
                    st.rerun()

            st.divider()
            for q in questions:
                q_id = q.get("id", "")
                q_text = q.get("text", "") or ""
                pd_q = q.get("productDetails", {}) or {}
                product = pd_q.get("productName", "") or q.get("productName", "") or ""
                created = q.get("createdDate", "")[:10] if q.get("createdDate") else ""
                preview_key_q = f"preview_q_{q_id}"

                with st.container():
                    col1, col2 = st.columns([6, 2])
                    with col1:
                        st.markdown(f'❓ <span style="font-size:12px;color:gray;">{product} · {created}</span>', unsafe_allow_html=True)
                        if q_text:
                            st.caption(f'"{q_text[:200]}{"..." if len(q_text)>200 else ""}"')
                    with col2:
                        pass

                    if q_id in auto_replied:
                        reply_text = auto_replied[q_id]
                        if "ИИ ошибка" in reply_text or "Error" in reply_text or "401" in reply_text:
                            del auto_replied[q_id]
                            save_json(_tmp(f"auto_replied_{idx}.json"), auto_replied)
                            st.warning("⚠️ Ошибочный ответ удалён")
                        else:
                            st.success("✅ Опубликован")
                            with st.expander("Посмотреть ответ"):
                                st.caption(reply_text)
                    elif preview_key_q in st.session_state:
                        preview_text = st.session_state[preview_key_q]
                        edited = st.text_area("✏️ Ответ ИИ — можно редактировать:", value=preview_text, key=f"edit_q_{q_id}", height=100)
                        c1, c2 = st.columns(2)
                        with c1:
                            if st.button("📤 Опубликовать", key=f"pub_q_{q_id}", use_container_width=True):
                                with st.spinner("Отправляется..."):
                                    time.sleep(1.1)
                                    ok = send_question_reply(fb_key, q_id, edited)
                                    if ok:
                                        auto_replied[q_id] = edited
                                        save_json(_tmp(f"auto_replied_{idx}.json"), auto_replied)
                                        del st.session_state[preview_key_q]
                                        st.rerun()
                                    else:
                                        st.error("Не отправлено")
                        with c2:
                            if st.button("🗑 Удалить", key=f"del_q_{q_id}", use_container_width=True):
                                del st.session_state[preview_key_q]
                                st.rerun()
                    else:
                        if st.button("🤖 Сгенерировать ответ ИИ", key=f"ai_q_{q_id}", use_container_width=True):
                            with st.spinner("ИИ генерирует ответ..."):
                                reply = ai_generate_reply(product, q_text, 5, "question")
                                st.session_state[preview_key_q] = reply
                                st.rerun()
                    st.divider()

def build_finance_excel(m, by_article, seb_data, logist_data, prod_rows, log_rows):
    import openpyxl as _oxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    buf = io.BytesIO()
    wb = _oxl.Workbook()
    ws = wb.active
    ws.title = "Финансы отчет"
    GREEN  = PatternFill("solid", fgColor="92D050")
    LGREEN = PatternFill("solid", fgColor="E2EFDA")
    BLUE   = PatternFill("solid", fgColor="BDD7EE")
    RED    = PatternFill("solid", fgColor="FF7F7F")
    BOLD   = Font(bold=True)
    BOLD12 = Font(bold=True, size=12)
    def bd():
        s = Side(style="thin"); return Border(left=s, right=s, top=s, bottom=s)
    def apply_bd(r1, r2, c1, c2):
        for r in range(r1, r2+1):
            for c in range(c1, c2+1):
                ws.cell(r, c).border = bd()
    row = 1
    ws.cell(row, 1, "НА ПЭЙ → ЧИСТАЯ ПРИБЫЛЬ").font = BOLD12
    row += 1
    napay_block = [
        ("авто", "К перечислению",          round(m["for_pay"]),       BLUE,  True),
        ("авто", "Логистика WB (доставка)", -round(m["logistic_wb"]),  None,  False),
        ("авто", "Хранение",                -round(m["storage"]),      None,  False),
        ("авто", "Штраф",                   -round(m["penalty"]),      None,  False),
        ("авто", "Операции на приёмке",     -round(m["priemka"]),      None,  False),
        ("авто", "Возврат × 2",             -round(m["vozvrat"]*2),    None,  False),
        ("авто", "НА ПЭЙ",                  round(m["napay"]),         BLUE,  True),
        ("авто", "Наш НДС",                 -round(m["nash_nds"]),     None,  False),
        ("авто", "ИПН 10%",                 -round(m["ipn"]),          None,  False),
        ("авто", "Себестоимость",           -round(m["tot_seb"]),      None,  False),
        ("авто", "Логистика до склада",     -round(m["logistika"]),    None,  False),
        ("авто", "Упаковка",                -round(m["upakovka"]),     None,  False),
        ("авто", "Реклама (удержания)",     -round(m["ads"]),          None,  False),
        ("қол",  "Самовыкуп",               -round(m["samovykup"]),    None,  False),
        ("қол",  "Бухгалтер",               -round(m["buhgalter"]),    None,  False),
        ("авто", "ЧИСТАЯ ПРИБЫЛЬ",          round(m["pribyl"]),        "profit", True),
        ("авто", "Рентабельность", f"{m['rent']*100:.1f}%",           "profit", False),
    ]
    start = row
    for tag, label, val, fill, bold in napay_block:
        c1 = ws.cell(row, 1, f"[{tag}] {label}")
        c2 = ws.cell(row, 2, val)
        actual = (GREEN if m["pribyl"] >= 0 else RED) if fill == "profit" else fill
        if actual:
            c1.fill = actual; c2.fill = actual
        if bold:
            c1.font = BOLD; c2.font = BOLD
        row += 1
    apply_bd(start, row-1, 1, 2)
    row += 2

    ws.cell(row, 1, "РАСЧЁТ НДС И ИПН").font = BOLD12
    row += 1
    nds_block = [
        ("вб услуги", round(m["vb_uslugi"])),
        ("вб реал тов − себ", round(m["vb_real_seb"])),
        ("база для ИПН", round(m["baza_ipn"])),
        ("ИПН 10%", round(m["ipn"])),
        ("общий НДС", round(m["obsh_nds"])),
        ("НДС поставщика", round(m["nds_post"])),
        ("НДС вб услуги", round(m["nds_vb"])),
        ("наш НДС", round(m["nash_nds"])),
    ]
    start = row
    for label, val in nds_block:
        ws.cell(row, 1, label); ws.cell(row, 2, val)
        row += 1
    apply_bd(start, row-1, 1, 2)
    row += 2

    ws.cell(row, 1, "ЧИСТАЯ ПРИБЫЛЬ ПО ТОВАРАМ").font = BOLD12
    row += 1
    headers = ["Артикул", "Продано", "вб получено", "расходы", "себес", "реклама", "прибыль", "%"]
    hr = row
    for col, h in enumerate(headers, 1):
        c = ws.cell(row, col, h); c.font = BOLD; c.fill = LGREEN
    row += 1
    for pr in prod_rows:
        ws.cell(row, 1, pr["Артикул"])
        ws.cell(row, 2, pr["Продано (шт)"])
        ws.cell(row, 3, pr["вб получено (₸)"])
        ws.cell(row, 4, pr["расходы (₸)"])
        ws.cell(row, 5, pr["себес (₸)"])
        ws.cell(row, 6, pr["реклама (₸)"])
        ws.cell(row, 7, pr["прибыль (₸)"])
        ws.cell(row, 8, pr["%"])
        if pr["прибыль (₸)"] < 0:
            ws.cell(row, 7).fill = RED
        row += 1
    apply_bd(hr, row-1, 1, 8)
    row += 2

    ws.cell(row, 1, "ЛОГИСТИКА ДО ВБ").font = BOLD12
    row += 1
    lheaders = ["Артикул", "Продано", "Короб ₸", "Шт/короб", "Итого ₸"]
    lr = row
    for col, h in enumerate(lheaders, 1):
        c = ws.cell(row, col, h); c.font = BOLD; c.fill = LGREEN
    row += 1
    for lg in log_rows:
        ws.cell(row, 1, lg["Артикул"])
        ws.cell(row, 2, lg["Продано (шт)"])
        ws.cell(row, 3, lg["Короб (₸)"])
        ws.cell(row, 4, lg["Шт/короб"])
        ws.cell(row, 5, lg["Итого (₸)"])
        row += 1
    apply_bd(lr, row-1, 1, 5)

    ws.column_dimensions["A"].width = 30
    for cl in ["B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[cl].width = 16
    wb.save(buf)
    return buf

def compute_finance(fin, seb_data, logist_data, ads_by_art, man):
    """Жаңа модель бойынша барлық қаржы көрсеткіштерін есептейді.
    Қайтарады: dict (жалпы көрсеткіштер + prod_rows + log_rows)."""
    r = 16/116
    for_pay    = fin.get("for_pay", 0)
    realizacia = fin.get("realizacia", 0)
    komissiya  = fin.get("komissiya", 0)
    vozmesh    = fin.get("vozmesh_pvz", 0)
    ads        = fin.get("ads", 0)
    storage    = fin.get("storage", 0)
    penalty    = fin.get("penalty", 0)
    logistic_wb = fin.get("logistic", 0)
    priemka    = fin.get("priemka", 0)
    vozvrat    = fin.get("vozvrat", 0)
    vozvrat_qty = fin.get("vozvrat_qty", 0)
    total_qty  = fin.get("total_qty", 0)
    by_article = fin.get("by_article", {})

    samovykup = man.get("samovykup", 0)
    buhgalter = man.get("buhgalter", 0)
    upak_price = man.get("upak_price", 100)  # 1 шт-қа упаковка бағасы (менеджер қояды)

    tot_seb = sum(seb_data.get(a, 0) * by_article.get(a, {}).get("qty", 0) for a in by_article)
    tot_qty_sold = sum(by_article.get(a, {}).get("qty", 0) for a in by_article)
    upakovka = tot_qty_sold * upak_price

    logistika = 0
    for art, d in by_article.items():
        qty_a = d.get("qty", 0)
        cfg = logist_data.get(art, {})
        box_price = cfg.get("box_price", 16000)
        box_qty   = cfg.get("box_qty", 0)
        if box_qty and box_qty > 0:
            logistika += (box_price / box_qty) * qty_a

    vb_uslugi   = komissiya + vozmesh + logistic_wb + storage + ads + priemka
    vb_real_seb = realizacia - logistic_wb - tot_seb
    baza_ipn    = vb_real_seb - vb_uslugi
    ipn         = baza_ipn * 0.10 if baza_ipn > 0 else 0
    obsh_nds    = realizacia * r
    nds_post    = tot_seb * r
    nds_vb      = vb_uslugi * r
    nash_nds    = obsh_nds - nds_post - nds_vb

    napay  = for_pay - logistic_wb - storage - penalty - priemka - vozvrat * 2 - ads
    pribyl = (napay - nash_nds - ipn - tot_seb
              - logistika - upakovka - samovykup - buhgalter)
    rent   = pribyl / napay if napay else 0

    obshie = (logistic_wb + storage + penalty + priemka + vozvrat*2
              + nash_nds + ipn + buhgalter + samovykup)
    per_sht = obshie / total_qty if total_qty else 0

    prod_rows = []
    log_rows  = []
    for art, d in by_article.items():
        qty_a = d.get("qty", 0)
        if qty_a <= 0:
            continue
        fp_a = d.get("for_pay", 0)
        seb_per = seb_data.get(art, 0)
        cfg = logist_data.get(art, {})
        box_price = cfg.get("box_price", 16000)
        box_qty   = cfg.get("box_qty", 0)
        logist_a  = (box_price / box_qty) * qty_a if box_qty and box_qty > 0 else 0
        vb_poluch = fp_a - per_sht * qty_a
        upak_a  = qty_a * upak_price
        rashody = upak_a + logist_a
        seb_tot_a = seb_per * qty_a
        if art in ads_by_art:
            reklama_a = ads_by_art.get(art, 0)
        else:
            reklama_a = ads * (fp_a / for_pay) if for_pay else 0
        profit_a = vb_poluch - rashody - seb_tot_a - reklama_a
        pct_a = profit_a / vb_poluch * 100 if vb_poluch else 0
        prod_rows.append({
            "Артикул": art,
            "Продано (шт)": qty_a,
            "вб получено (₸)": round(vb_poluch),
            "расходы (₸)": round(rashody),
            "себес (₸)": round(seb_tot_a),
            "реклама (₸)": reklama_a,
            "прибыль (₸)": round(profit_a),
            "%": round(pct_a, 1),
        })
        log_rows.append({
            "Артикул": art,
            "Продано (шт)": qty_a,
            "Короб (₸)": box_price,
            "Шт/короб": box_qty,
            "Итого (₸)": round(logist_a),
        })

    return {
        "for_pay": for_pay, "realizacia": realizacia, "ads": ads,
        "storage": storage, "penalty": penalty, "logistic_wb": logistic_wb,
        "priemka": priemka, "vozvrat": vozvrat, "vozvrat_qty": vozvrat_qty,
        "total_qty": total_qty, "tot_seb": tot_seb, "tot_qty_sold": tot_qty_sold,
        "upakovka": upakovka, "logistika": logistika, "samovykup": samovykup,
        "upak_price": upak_price,
        "buhgalter": buhgalter, "vb_uslugi": vb_uslugi, "vb_real_seb": vb_real_seb,
        "baza_ipn": baza_ipn, "ipn": ipn, "obsh_nds": obsh_nds, "nds_post": nds_post,
        "nds_vb": nds_vb, "nash_nds": nash_nds, "napay": napay, "pribyl": pribyl,
        "rent": rent, "per_sht": per_sht,
        "prod_rows": prod_rows, "log_rows": log_rows,
    }

def show_finance_tab(store, df):
    idx = store["idx"]
    stats_key = store["stats_key"]
    finance_key = store["finance_key"]
    name = store["name"]
    role = st.session_state.get("role", "manager")

    st.markdown("#### 💰 Финансовый отчёт")
    col1, col2, col3 = st.columns([1, 1, 1])
    with col1:
        date_from = st.date_input("Начало периода", value=date.today()-timedelta(days=7), key=f"fin_from_{idx}")
    with col2:
        date_to = st.date_input("Конец периода", value=date.today(), key=f"fin_to_{idx}")
    with col3:
        st.markdown("<br>", unsafe_allow_html=True)
        load_fin = st.button("🔄 Загрузить финансы", key=f"fin_load_{idx}", use_container_width=True)

    fin_key = f"finance_{idx}"
    use_key = finance_key if finance_key else stats_key
    period_key = f"fin_period_{idx}"
    current_period = f"{date_from}_{date_to}"
    if st.session_state.get(period_key) != current_period:
        if fin_key in st.session_state:
            del st.session_state[fin_key]
        st.session_state[period_key] = current_period

    if load_fin:
        with st.spinner(f"[{name}] Загрузка финансового отчёта..."):
            try:
                rows = fetch_report_detail(use_key, date_from.strftime("%Y-%m-%d"), date_to.strftime("%Y-%m-%d"), store_name=name)
                fin = parse_finance(rows)
                st.session_state[fin_key] = fin
                st.success("✅ Загружено!")
            except Exception as e:
                st.error(f"Ошибка: {e}")

    if fin_key not in st.session_state:
        st.info("👆 Выберите период и нажмите **«Загрузить финансы»**")
        return

    fin = st.session_state[fin_key]

    all_upak = load_json(SEBEST_FILE)
    saved_upak = all_upak.get(f"_upak_{idx}", 100)
    man_key = f"fin_manual_{idx}"
    if man_key not in st.session_state:
        st.session_state[man_key] = {"samovykup": 0, "buhgalter": 0, "upak_price": saved_upak}
    man = st.session_state[man_key]
    if "upak_price" not in man:
        man["upak_price"] = saved_upak

    ads_key = f"ads_by_art_{idx}"
    if ads_key not in st.session_state:
        st.session_state[ads_key] = {}
    ads_by_art = st.session_state[ads_key]

    all_seb = load_json(SEBEST_FILE)
    seb_data = all_seb.get(str(idx), {})
    all_log = load_json(LOGISTIC_FILE)
    logist_data = all_log.get(str(idx), {})

    m = compute_finance(fin, seb_data, logist_data, ads_by_art, man)

    def fmt(n): return f"{round(n):,} ₸".replace(",", " ")
    def fmtN(n): return f"{round(n):,}".replace(",", " ")

    c1, c2, c3 = st.columns(3)
    c1.metric("К перечислению", fmt(m["for_pay"]))
    c2.metric("На пэй", fmt(m["napay"]))
    c3.metric("Чистая прибыль", fmt(m["pribyl"]), delta=f"{m['rent']*100:.1f}%")

    st.divider()
    col_left, col_right = st.columns(2)

    with col_left:
        st.markdown("**🧮 На пэй → Чистая прибыль**")
        def rl(label, value, color="red"):
            r1, r2 = st.columns([3, 2])
            r1.caption(label)
            clr = "#185FA5" if color=="blue" else ("#A32D2D" if color=="red" else "#3B6D11")
            r2.markdown(f"<p style='text-align:right;color:{clr};font-weight:500;'>{value}</p>", unsafe_allow_html=True)
        rl("[авто] К перечислению", fmt(m["for_pay"]), "blue")
        rl("[авто] Логистика WB (доставка)", f"- {fmt(m['logistic_wb'])}")
        rl("[авто] Хранение", f"- {fmt(m['storage'])}")
        rl("[авто] Штраф", f"- {fmt(m['penalty'])}")
        rl("[авто] Операции на приёмке", f"- {fmt(m['priemka'])}")
        rl(f"[авто] Возврат × 2 ({fmtN(m['vozvrat_qty'])} шт)", f"- {fmt(m['vozvrat']*2)}")
        st.markdown(f"**На пэй: :blue[{fmt(m['napay'])}]**")
        st.divider()
        rl("[авто] Наш НДС", f"- {fmt(m['nash_nds'])}")
        rl("[авто] ИПН 10%", f"- {fmt(m['ipn'])}")
        rl("[авто] Себестоимость", f"- {fmt(m['tot_seb'])}")
        rl("[авто] Логистика до склада", f"- {fmt(m['logistika'])}")
        rl(f"[авто] Упаковка ({fmtN(m['tot_qty_sold'])} × {m['upak_price']:.0f}₸)", f"- {fmt(m['upakovka'])}")
        rl("[авто] Реклама (удержания)", f"- {fmt(m['ads'])}")
        if role in ("manager", "submanager"):
            new_upak = st.number_input("[қол] Упаковка 1 шт (₸)", value=float(man.get("upak_price", 100)), min_value=0.0, step=5.0, key=f"upak_{idx}",
                                       help="Әр ИП-де упаковка бағасы әртүрлі — 1 данаға қанша төлейтініңізді жазыңыз")
            new_samo = st.number_input("[қол] Самовыкуп (₸)", value=float(man["samovykup"]), min_value=0.0, step=1000.0, key=f"samo_{idx}")
            new_buh = st.number_input("[қол] Бухгалтер (₸)", value=float(man["buhgalter"]), min_value=0.0, step=1000.0, key=f"buh_{idx}")
            if (new_samo != man["samovykup"] or new_buh != man["buhgalter"]
                    or new_upak != man.get("upak_price", 100)):
                st.session_state[man_key] = {"samovykup": new_samo, "buhgalter": new_buh, "upak_price": new_upak}
                all_upak[f"_upak_{idx}"] = new_upak
                save_json(SEBEST_FILE, all_upak)
                st.rerun()
        else:
            rl(f"[қол] Упаковка 1 шт", f"{m['upak_price']:.0f} ₸", "plain")
            rl("[қол] Самовыкуп", f"- {fmt(m['samovykup'])}", "plain")
            rl("[қол] Бухгалтер", f"- {fmt(m['buhgalter'])}", "plain")
        st.divider()
        st.markdown(f"### Чистая прибыль: :green[{fmt(m['pribyl'])}]")
        st.caption(f"Рентабельность: {m['rent']*100:.1f}%")

    with col_right:
        st.markdown("**🧾 Расчёт НДС и ИПН**")
        st.markdown(f"- вб услуги: {fmt(m['vb_uslugi'])}")
        st.caption("комиссия + возмещение ПВЗ + доставка + хранение + реклама + приёмка")
        st.markdown(f"- вб реал тов − себ: {fmt(m['vb_real_seb'])}")
        st.caption("реализация − доставка − себестоимость")
        st.markdown(f"- **база для ИПН: {fmt(m['baza_ipn'])}**")
        st.markdown(f"- ИПН 10%: :red[{fmt(m['ipn'])}]")
        st.divider()
        st.markdown(f"- общий НДС: {fmt(m['obsh_nds'])}")
        st.markdown(f"- НДС поставщика: {fmt(m['nds_post'])}")
        st.markdown(f"- НДС вб услуги: {fmt(m['nds_vb'])}")
        st.markdown(f"- **наш НДС: :red[-{fmt(m['nash_nds'])}]**")
        st.divider()
        st.markdown("**📊 Продажи**")
        st.markdown(f"- Реализация (Пр): {fmt(m['realizacia'])}")
        st.markdown(f"- Всего продано: **{fmtN(m['total_qty'])} шт**")
        st.markdown(f"- Возврат: :red[**{fmtN(m['vozvrat_qty'])} шт**]")
        st.caption(f"1 шт-қа жалпы шығын: {fmt(m['per_sht'])}")

    st.divider()

    buf = build_finance_excel(m, fin.get("by_article", {}), seb_data, logist_data, m["prod_rows"], m["log_rows"])
    period_str = f"{date_from.strftime('%d.%m')}-{date_to.strftime('%d.%m.%Y')}"
    st.download_button(f"⬇️ Скачать Excel — {store['name']} ({period_str})",
        data=buf.getvalue(), file_name=f"Финансы_{store['name']}_{period_str}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"fin_dl_{idx}")

    st.divider()

    if m["prod_rows"]:
        with st.expander("📦 Чистая прибыль по товарам", expanded=True):
            prod_df = pd.DataFrame(m["prod_rows"])
            def sp(val):
                if pd.isna(val): return ""
                return "color:#3B6D11;font-weight:bold" if val>=0 else "color:#A32D2D;font-weight:bold"
            styled = prod_df.style.map(sp, subset=["прибыль (₸)", "%"])
            st.dataframe(styled, use_container_width=True, height=400,
                column_config={
                    "вб получено (₸)": st.column_config.NumberColumn(format="%d ₸"),
                    "расходы (₸)": st.column_config.NumberColumn(format="%d ₸"),
                    "себес (₸)": st.column_config.NumberColumn(format="%d ₸"),
                    "реклама (₸)": st.column_config.NumberColumn(format="%d ₸"),
                    "прибыль (₸)": st.column_config.NumberColumn(format="%d ₸"),
                    "%": st.column_config.NumberColumn(format="%.1f%%"),
                })
            if role in ("manager", "submanager"):
                st.divider()
                st.caption("✏️ Реклама — введите вручную по товару")
                arts_list = [r["Артикул"] for r in m["prod_rows"]]
                ads_tbl = pd.DataFrame([{"Артикул": a, "Реклама (₸)": ads_by_art.get(a, 0)} for a in arts_list])
                ads_edited = st.data_editor(ads_tbl, use_container_width=True, height=300, key=f"ads_editor_{idx}",
                    column_config={
                        "Артикул": st.column_config.TextColumn(disabled=True),
                        "Реклама (₸)": st.column_config.NumberColumn(format="%d ₸", min_value=0),
                    })
                new_ads = dict(zip(ads_edited["Артикул"], ads_edited["Реклама (₸)"]))
                if new_ads != ads_by_art:
                    st.session_state[ads_key] = new_ads
                    st.rerun()

    st.divider()

    cL, cR = st.columns(2)
    with cL:
        with st.expander("🗂️ Себестоимость — по товарам", expanded=False):
            if m["prod_rows"]:
                seb_rows = [{"Артикул": r["Артикул"], "Продано (шт)": r["Продано (шт)"],
                             "Себест/шт (₸)": seb_data.get(r["Артикул"], 0),
                             "Итого (₸)": seb_data.get(r["Артикул"], 0)*r["Продано (шт)"]}
                            for r in m["prod_rows"]]
                seb_df = pd.DataFrame(seb_rows)
                if role in ("manager", "submanager"):
                    st.caption("✏️ Себест/шт өзгертіп Enter басыңыз")
                    edited_seb = st.data_editor(seb_df, use_container_width=True, height=350,
                        column_config={
                            "Артикул": st.column_config.TextColumn(disabled=True),
                            "Продано (шт)": st.column_config.NumberColumn(format="%d шт", disabled=True),
                            "Себест/шт (₸)": st.column_config.NumberColumn(format="%d ₸", min_value=0),
                            "Итого (₸)": st.column_config.NumberColumn(format="%d ₸", disabled=True),
                        })
                    new_seb = dict(zip(edited_seb["Артикул"], edited_seb["Себест/шт (₸)"]))
                    if new_seb != {k: seb_data.get(k, 0) for k in new_seb}:
                        merged = dict(seb_data); merged.update(new_seb)
                        all_seb[str(idx)] = merged
                        save_json(SEBEST_FILE, all_seb)
                        st.rerun()
                else:
                    st.dataframe(seb_df, use_container_width=True, height=350)

    with cR:
        with st.expander("🚚 Логистика до ВБ", expanded=False):
            if m["log_rows"]:
                log_df = pd.DataFrame([{"Артикул": r["Артикул"], "Продано (шт)": r["Продано (шт)"],
                                        "Короб (₸)": r["Короб (₸)"], "Шт/короб": r["Шт/короб"],
                                        "Итого (₸)": r["Итого (₸)"]} for r in m["log_rows"]])
                if role in ("manager", "submanager"):
                    st.caption("✏️ Короб ₸ мен Шт/короб өзгертіңіз — итого авто")
                    edited_log = st.data_editor(log_df, use_container_width=True, height=350, key=f"log_editor_{idx}",
                        column_config={
                            "Артикул": st.column_config.TextColumn(disabled=True),
                            "Продано (шт)": st.column_config.NumberColumn(format="%d шт", disabled=True),
                            "Короб (₸)": st.column_config.NumberColumn(format="%d ₸", min_value=0),
                            "Шт/короб": st.column_config.NumberColumn(format="%d", min_value=0),
                            "Итого (₸)": st.column_config.NumberColumn(format="%d ₸", disabled=True),
                        })
                    new_log = {}
                    for _, rr in edited_log.iterrows():
                        new_log[rr["Артикул"]] = {"box_price": int(rr["Короб (₸)"]), "box_qty": int(rr["Шт/короб"])}
                    if new_log != {k: logist_data.get(k, {}) for k in new_log}:
                        merged = dict(logist_data); merged.update(new_log)
                        all_log[str(idx)] = merged
                        save_json(LOGISTIC_FILE, all_log)
                        st.rerun()
                    st.caption(f"Итого → Логистика до склада: {fmt(m['logistika'])}")
                else:
                    st.dataframe(log_df, use_container_width=True, height=350)

def show_store(store, df, sales30, filter_status, search):
    idx = store["idx"]
    role = st.session_state.get("role", "manager")

    if df.empty:
        st.warning("Нет данных или не загружено")
        return

    tab_ostatok, tab_fbs, tab_finance, tab_sales, tab_feedback = st.tabs([
        "📦 Остатки", "🚚 FBS Заказы", "💰 Финансы", "📅 Ежедневный отчёт", "💬 Отзывы & Вопросы"
    ])

    with tab_fbs:
        show_fbs_tab(store)

    with tab_finance:
        show_finance_tab(store, df)

    with tab_sales:
        show_sales_returns_tab(store)

    with tab_feedback:
        show_feedback_tab(store)

    with tab_ostatok:
        dff = df.copy()
        if filter_status == "Ноль":
            dff = dff[dff["qty"] == 0]
        elif filter_status == "Мало (1–200)":
            dff = dff[(dff["qty"] >= 1) & (dff["qty"] <= 200)]
        elif filter_status == "Хорошо (201–500)":
            dff = dff[(dff["qty"] > 200) & (dff["qty"] <= 500)]
        elif filter_status == "Достаточно (500+)":
            dff = dff[dff["qty"] > 500]
        if search:
            dff = dff[dff["supplierArticle"].astype(str).str.contains(search, case=False, na=False)]

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("📦 Общий остаток", f"{int(df['total'].sum()):,} шт".replace(",", " "))
        c2.metric("📊 Позиций", len(df))
        c3.metric("🔴 Критические", int(df["turnover"].dropna().apply(lambda x: x <= 10).sum()))
        c4.metric("⚫ Нулевой остаток", int((df["qty"] == 0).sum()))
        st.divider()
        st.markdown("#### 📊 Итоговый отчёт")

        all_fbo = load_json(FBO_FILE)
        fbo_data = all_fbo.get(str(idx), {})

        result = dff[["supplierArticle", "qty", "in_way_client", "daily_avg", "status"]].copy()
        result["FBO в пути"] = result["supplierArticle"].map(fbo_data).fillna(0).astype(int)
        result["Общий остаток"] = result["qty"] + result["in_way_client"] + result["FBO в пути"]
        result["Оборачиваемость"] = result.apply(
            lambda r: round(r["Общий остаток"]/r["daily_avg"]) if r["daily_avg"]>0 else None, axis=1)
        result = result.rename(columns={
            "supplierArticle": "Артикул", "qty": "Остаток",
            "in_way_client": "В пути к клиенту", "daily_avg": "Ср. продаж/день", "status": "Статус"})
        result = result[["Артикул", "Остаток", "В пути к клиенту", "FBO в пути",
                         "Общий остаток", "Ср. продаж/день", "Оборачиваемость", "Статус"]]

        def style_turn(val):
            if pd.isna(val): return ""
            return "background-color:#FCEBEB;color:#A32D2D;font-weight:bold" if val<=10 else ""
        def style_stat(val):
            mm = {"Ноль":"background-color:#FCEBEB;color:#A32D2D","Мало":"background-color:#FAEEDA;color:#854F0B",
                  "Хорошо":"background-color:#EAF3DE;color:#3B6D11","Достаточно":"background-color:#E6F1FB;color:#185FA5"}
            return mm.get(val, "")

        styled = result.style.map(style_turn, subset=["Оборачиваемость"]).map(style_stat, subset=["Статус"])
        st.dataframe(styled, use_container_width=True, height=460,
            column_config={
                "Остаток": st.column_config.NumberColumn(format="%d шт"),
                "В пути к клиенту": st.column_config.NumberColumn(format="%d шт"),
                "FBO в пути": st.column_config.NumberColumn(format="%d шт"),
                "Общий остаток": st.column_config.NumberColumn(format="%d шт"),
                "Ср. продаж/день": st.column_config.NumberColumn(format="%.1f"),
                "Оборачиваемость": st.column_config.NumberColumn(format="%d дн"),
            })
        st.caption(f"Показано: {len(result)} позиций")

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            pd.DataFrame().to_excel(writer, sheet_name="_tmp", index=False)
            result.to_excel(writer, index=False, sheet_name="Остатки WB")
        st.download_button(f"⬇️ Excel — {store['name']}", data=buf.getvalue(),
            file_name=f"WB_{store['name']}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_{idx}")

        if role in ("manager", "submanager"):
            st.divider()
            st.markdown("#### ✏️ FBO в пути")
            fbo_tbl = dff[["supplierArticle"]].copy()
            fbo_tbl.columns = ["Артикул"]
            fbo_tbl["FBO в пути"] = fbo_tbl["Артикул"].map(fbo_data).fillna(0).astype(int)
            fbo_edited = st.data_editor(fbo_tbl, use_container_width=True, height=400, key=f"fbo_editor_{idx}",
                column_config={
                    "Артикул": st.column_config.TextColumn(disabled=True),
                    "FBO в пути": st.column_config.NumberColumn(format="%d шт", min_value=0),
                })
            new_fbo = dict(zip(fbo_edited["Артикул"], fbo_edited["FBO в пути"]))
            if new_fbo != fbo_data:
                all_fbo[str(idx)] = new_fbo
                save_json(FBO_FILE, all_fbo)
                st.rerun()

# ── SIDEBAR ──
stores = get_stores()
_role = st.session_state.get("role", "manager")
_store_access = st.session_state.get("store_access", None)
if _role == "owner" and _store_access:
    visible_stores = [s for s in stores if s["idx"] == _store_access]
elif _role == "submanager" and _store_access:
    visible_stores = [s for s in stores if s["idx"] in _store_access]
else:
    visible_stores = stores

with st.sidebar:
    st.header("⚙️ Настройки")
    if not visible_stores:
        st.warning("Нет магазинов!")
    fetch_btn = st.button("🔄 Загрузить все", use_container_width=True)
    st.divider()

    _role_now = st.session_state.get("role", "manager")
    if _role_now in ("manager", "submanager"):
        _cur_view = st.session_state.get("nav_view", "store_0")
        _mg_on = st.session_state.get("show_magkein", False)
        _store_names = [s["name"] for s in visible_stores]
        try:
            _cur_idx = int(_cur_view.split("_")[1])
        except:
            _cur_idx = 0
        _cur_store_name = visible_stores[_cur_idx]["name"] if not _mg_on and 0 <= _cur_idx < len(visible_stores) else _store_names[0]
        _nav_sel = st.selectbox("ВБ Кабинеты", _store_names,
            index=_store_names.index(_cur_store_name) if _cur_store_name in _store_names else 0, key="wb_nav_select")
        _prev_sel = st.session_state.get("wb_prev_sel", _store_names[0])
        if _nav_sel != _prev_sel:
            st.session_state.wb_prev_sel = _nav_sel
            for _ci, _cs in enumerate(visible_stores):
                if _cs["name"] == _nav_sel:
                    st.session_state.nav_view = f"store_{_ci}"
                    st.session_state.show_magkein = False
                    st.rerun()
        st.session_state.wb_prev_sel = _nav_sel
        if _role_now == "manager":
            st.divider()
            _mg_label = "✅ MAGKEIN — закрыть" if _mg_on else "📊 Отчёт MAGKEIN"
            if st.button(_mg_label, key="mg_toggle_btn", use_container_width=True):
                st.session_state.show_magkein = not _mg_on
                st.rerun()
    else:
        for _s in visible_stores:
            st.markdown(f"**{_s['name']}**")

    st.divider()
    search = st.text_input("🔍 Поиск по артикулу")
    st.markdown("""<div style='font-size:11px;color:#854F0B;background:#FAEEDA;padding:8px;border-radius:6px;margin-top:8px;'>
    🔴 Красная ячейка = оборачиваемость ≤ 10 дней</div>""", unsafe_allow_html=True)
    if st.button("🚪 Выйти"):
        st.session_state.role = None
        st.session_state.store_access = None
        st.query_params.clear()
        st.rerun()

# ── НЕГІЗГІ ──
st.title("📦 Wildberries Отчёт")
st.caption(f"Обновлено: {datetime.now().strftime('%d.%m.%Y %H:%M')}")

if not visible_stores:
    st.info("👈 Добавьте данные магазинов в Secrets")
    st.stop()

if fetch_btn:
    for s in visible_stores:
        df, sales30, errors = load_store_data(s)
        st.session_state[f"df_{s['idx']}"] = df
        st.session_state[f"sales30_{s['idx']}"] = sales30
        for e in errors:
            st.warning(f"[{s['name']}] ⚠️ {e}")
    st.success("✅ Все магазины загружены!")
    st.rerun()

_show_magkein = (st.session_state.get("role", "manager") == "manager"
                 and st.session_state.get("show_magkein", False))

if _show_magkein:
    st.markdown('## 📊 MAGKEIN — Общий финансовый отчёт')
    st.markdown('Объединённые результаты всех магазинов')
    st.divider()

    col1, col2, col3 = st.columns([1, 1, 1])
    with col1:
        mg_date_from = st.date_input("Начало периода", value=date.today()-timedelta(days=7), key="mg_from")
    with col2:
        mg_date_to = st.date_input("Конец периода", value=date.today(), key="mg_to")
    with col3:
        st.markdown("<br>", unsafe_allow_html=True)
        mg_load = st.button("🔄 Загрузить все", key="mg_load", use_container_width=True)

    mg_key = "magkein_data"
    mg_period_key = "magkein_period"
    current_mg_period = f"{mg_date_from}_{mg_date_to}"
    if st.session_state.get(mg_period_key) != current_mg_period:
        if mg_key in st.session_state:
            del st.session_state[mg_key]
        st.session_state[mg_period_key] = current_mg_period

    if mg_load:
        mg_results = {}
        for s in visible_stores:
            use_key = s["finance_key"] if s["finance_key"] else s["stats_key"]
            with st.spinner(f"[{s['name']}] загружается..."):
                try:
                    rows = fetch_report_detail(use_key, mg_date_from.strftime("%Y-%m-%d"), mg_date_to.strftime("%Y-%m-%d"), store_name=s["name"])
                    fin = parse_finance(rows)
                    mg_results[s["name"]] = fin
                except Exception as e:
                    st.error(f"[{s['name']}] ошибка: {e}")
                    mg_results[s["name"]] = {}
        st.session_state[mg_key] = mg_results
        st.rerun()

    if mg_key not in st.session_state:
        st.info("👆 Период таңдап **«Загрузить все»** нажмите кнопку")
    else:
        mg_results = st.session_state[mg_key]
        all_seb = load_json(SEBEST_FILE)
        all_log = load_json(LOGISTIC_FILE)

        def fmt(n): return f"{round(n):,} ₸".replace(",", " ")
        def fmtN(n): return f"{round(n):,}".replace(",", " ")

        summary_rows = []
        total_for_pay = total_napay = total_profit = total_qty = total_vozvrat_qty = 0

        for s in visible_stores:
            fin = mg_results.get(s["name"], {})
            if not fin:
                continue
            idx = s["idx"]
            seb_data = all_seb.get(str(idx), {})
            logist_data = all_log.get(str(idx), {})
            man = st.session_state.get(f"fin_manual_{idx}", {"samovykup": 0, "buhgalter": 0})
            if "upak_price" not in man:
                man = dict(man)
                man["upak_price"] = all_seb.get(f"_upak_{idx}", 100)
            ads_by_art = st.session_state.get(f"ads_by_art_{idx}", {})
            m = compute_finance(fin, seb_data, logist_data, ads_by_art, man)
            summary_rows.append({
                "Магазин": s["name"],
                "К перечислению (₸)": round(m["for_pay"]),
                "На пэй (₸)": round(m["napay"]),
                "Себестоимость (₸)": round(m["tot_seb"]),
                "Продано (шт)": m["total_qty"],
                "Возврат (шт)": m["vozvrat_qty"],
                "Чистая прибыль (₸)": round(m["pribyl"]),
                "Рентабельность (%)": round(m["rent"]*100, 1),
            })
            total_for_pay += m["for_pay"]
            total_napay += m["napay"]
            total_profit += m["pribyl"]
            total_qty += m["total_qty"]
            total_vozvrat_qty += m["vozvrat_qty"]

        if not summary_rows:
            st.warning("Нет данных")
        else:
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("💰 К перечислению", fmt(total_for_pay))
            c2.metric("📊 На пэй (жалпы)", fmt(total_napay))
            c3.metric("✅ Таза пайда (жалпы)", fmt(total_profit),
                      delta=f"{total_profit/total_for_pay*100:.1f}%" if total_for_pay>0 else "0%")
            c4.metric("📦 Сатылды (жалпы)", f"{fmtN(total_qty)} шт")
            st.divider()
            st.markdown("#### 🏪 Сравнение по магазинам")
            summary_rows.append({
                "Магазин": "🔷 ИТОГО (MAGKEIN)",
                "К перечислению (₸)": round(total_for_pay),
                "На пэй (₸)": round(total_napay),
                "Себестоимость (₸)": sum(r["Себестоимость (₸)"] for r in summary_rows),
                "Продано (шт)": total_qty,
                "Возврат (шт)": total_vozvrat_qty,
                "Чистая прибыль (₸)": round(total_profit),
                "Рентабельность (%)": round(total_profit/total_for_pay*100, 1) if total_for_pay>0 else 0,
            })
            mg_df = pd.DataFrame(summary_rows)
            def style_mg(row):
                if row["Магазин"].startswith("🔷"):
                    return ["font-weight:bold; background-color:#E6F1FB"] * len(row)
                pv = row["Чистая прибыль (₸)"]
                if isinstance(pv, (int, float)) and pv < 0:
                    return ["background-color:#FCEBEB"] * len(row)
                return [""] * len(row)
            styled_mg = mg_df.style.apply(style_mg, axis=1)
            st.dataframe(styled_mg, use_container_width=True, height=200,
                column_config={
                    "К перечислению (₸)": st.column_config.NumberColumn(format="%d ₸"),
                    "На пэй (₸)": st.column_config.NumberColumn(format="%d ₸"),
                    "Себестоимость (₸)": st.column_config.NumberColumn(format="%d ₸"),
                    "Продано (шт)": st.column_config.NumberColumn(format="%d шт"),
                    "Возврат (шт)": st.column_config.NumberColumn(format="%d шт"),
                    "Чистая прибыль (₸)": st.column_config.NumberColumn(format="%d ₸"),
                    "Рентабельность (%)": st.column_config.NumberColumn(format="%.1f%%"),
                })
            st.divider()
            st.markdown("#### 📊 Чистая прибыль по магазинам")
            chart_df = mg_df[mg_df["Магазин"] != "🔷 ИТОГО (MAGKEIN)"][["Магазин", "Чистая прибыль (₸)"]].set_index("Магазин")
            st.bar_chart(chart_df, height=300)

            st.divider()
            st.markdown("#### 📦 Продажи по артикулам (все магазины)")
            art_combined = {}
            for s in visible_stores:
                fin = mg_results.get(s["name"], {})
                for art, data in fin.get("by_article", {}).items():
                    art_combined.setdefault(art, {})[s["name"]] = data.get("qty", 0)
            if art_combined:
                store_names_list = [s["name"] for s in visible_stores if mg_results.get(s["name"])]
                art_rows = []
                for art, sq in art_combined.items():
                    row = {"Артикул": art}
                    for sn in store_names_list:
                        row[sn] = sq.get(sn, 0)
                    row["ИТОГО (шт)"] = sum(sq.values())
                    art_rows.append(row)
                art_df = pd.DataFrame(art_rows).sort_values("ИТОГО (шт)", ascending=False).reset_index(drop=True)
                def style_art(row):
                    if row["ИТОГО (шт)"] == art_df["ИТОГО (шт)"].max():
                        return ["background-color:#EAF3DE"] * len(row)
                    return [""] * len(row)
                col_cfg = {"ИТОГО (шт)": st.column_config.NumberColumn(format="%d шт")}
                for sn in store_names_list:
                    col_cfg[sn] = st.column_config.NumberColumn(format="%d шт")
                st.dataframe(art_df.style.apply(style_art, axis=1), use_container_width=True,
                    height=min(400, 40+len(art_df)*35), column_config=col_cfg)

            st.divider()
            buf_mg = io.BytesIO()
            import openpyxl as _oxl2
            from openpyxl.styles import Font as _Font2, PatternFill as _Fill2
            _wb_mg = _oxl2.Workbook()
            _ws_mg = _wb_mg.active
            _ws_mg.title = "MAGKEIN отчет"
            headers_mg = ["Магазин", "К перечислению (₸)", "На пэй (₸)", "Себестоимость (₸)",
                          "Продано (шт)", "Возврат (шт)", "Чистая прибыль (₸)", "Рентабельность (%)"]
            for col, h in enumerate(headers_mg, 1):
                c = _ws_mg.cell(1, col, h)
                c.font = _Font2(bold=True)
                c.fill = _Fill2("solid", fgColor="BDD7EE")
            for r_idx, row in enumerate(summary_rows, 2):
                for col, h in enumerate(headers_mg, 1):
                    cell = _ws_mg.cell(r_idx, col, row[h])
                    if row["Магазин"].startswith("🔷"):
                        cell.font = _Font2(bold=True)
                        cell.fill = _Fill2("solid", fgColor="92D050")
            for cl in ["A","B","C","D","E","F","G","H"]:
                _ws_mg.column_dimensions[cl].width = 22
            _wb_mg.save(buf_mg)
            period_str = f"{mg_date_from.strftime('%d.%m')}-{mg_date_to.strftime('%d.%m.%Y')}"
            st.download_button(f"⬇️ Excel — MAGKEIN ({period_str})", data=buf_mg.getvalue(),
                file_name=f"MAGKEIN_{period_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="mg_dl")

else:
    _nav = st.session_state.get("nav_view", "store_0")
    try:
        _nav_idx = int(_nav.split("_")[1])
    except:
        _nav_idx = 0
    _nav_idx = min(_nav_idx, len(visible_stores) - 1)
    _store = visible_stores[_nav_idx]
    st.markdown(f"### 🏪 {_store['name']}")
    st.divider()
    _df_key = f"df_{_store['idx']}"
    if _df_key not in st.session_state or st.session_state[_df_key] is None:
        st.info("👈 Нажмите **«Загрузить все»**")
    else:
        _sales30 = st.session_state.get(f"sales30_{_store['idx']}", pd.DataFrame())
        show_store(_store, st.session_state[_df_key], _sales30, "Все", search)
        
